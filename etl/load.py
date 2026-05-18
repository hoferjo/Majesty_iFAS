from etl.transform import getDocPath
def export_docs_xlsx(article_list, output_path=None, template_xlsx_path=None, cache_csv_path=None):
    """
    For each article with a drawing number, find the doc file and write a row to XLSX.
    Columns: artnr, zeichnungsnummer, zeichnungsindex, bezeichnung, doc_path, doc_filename
    Returns the XLSX file path.
    """
    # Instead of regenerating, read from the DMSDocuments cache CSV.
    import csv
    from copy import copy

    if output_path is None:
        output_path = BASE_DIR / "data" / "processed" / "docs_export.xlsx"

    def _normalize_header(value):
        import unicodedata

        normalized = unicodedata.normalize("NFKD", str(value or ""))
        return "".join(ch for ch in normalized if ch.isalnum()).lower()

    if template_xlsx_path:
        template_xlsx_path = Path(template_xlsx_path)
        if not template_xlsx_path.exists():
            raise FileNotFoundError(f"Template not found at {template_xlsx_path}")

        dms_cache_path = Path(cache_csv_path) if cache_csv_path else BASE_DIR / "data" / "processed" / "cache" / "sheets" / "DMSDocuments.csv"
        if not dms_cache_path.exists():
            raise FileNotFoundError(f"DMSDocuments cache CSV not found at {dms_cache_path}")

        wb = openpyxl.load_workbook(template_xlsx_path)
        if "DMSDokumente" not in wb.sheetnames:
            raise ValueError("DMSDokumente sheet not found in DMS template")

        ws = wb["DMSDokumente"]
        header_row = 7
        data_start_row = header_row + 1
        max_col = ws.max_column
        template_styles = [copy(ws.cell(row=data_start_row, column=col_idx)._style) for col_idx in range(1, max_col + 1)]

        # Remove the sample row(s) from the template before writing the new data.
        for row_idx in range(data_start_row, ws.max_row + 1):
            for col_idx in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

        with open(dms_cache_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=';')
            rows = list(reader)
            cache_headers = list(reader.fieldnames or [])

        if not rows:
            raise ValueError("No rows found in DMSDocuments cache CSV.")

        cache_header_map = {_normalize_header(header): header for header in cache_headers if str(header or "").strip()}
        column_map = {
            2: "Artikelnummer",
            3: "Artikelnummer",
            4: "Identifikation",
            5: "DMSDocId",
            6: "Bezeichnung",
            7: "Kategorie",
            8: "Dokumenten Speicherort",
            9: "Gleiches Dokument ersetzen",
            10: "Als Verknüpfung importieren",
            11: "Organisationseinheit",
        }

        def _resolve_cache_key(template_label):
            normalized_label = _normalize_header(template_label)
            if normalized_label == _normalize_header("Als Verknüpfung importieren"):
                return cache_header_map.get(normalized_label) or cache_header_map.get(_normalize_header("ImportAsShortcutLink"))
            return cache_header_map.get(normalized_label)

        for row_offset, row in enumerate(rows, start=data_start_row):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_offset, column=col_idx)
                if col_idx <= len(template_styles):
                    cell._style = copy(template_styles[col_idx - 1])
                template_label = column_map.get(col_idx)
                value = ""
                if template_label:
                    cache_key = _resolve_cache_key(template_label)
                    if cache_key:
                        value = row.get(cache_key, "")
                cell.value = "" if value is None else str(value)
                cell.number_format = "@"

        wb.save(output_path)
        return str(output_path)

    # Legacy fallback: export the cache CSV directly into a simple workbook.
    dms_cache_path = BASE_DIR / "data" / "processed" / "cache" / "sheets" / "DMSDocuments.csv"
    if not dms_cache_path.exists():
        raise FileNotFoundError(f"DMSDocuments cache CSV not found at {dms_cache_path}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docs"
    # Use the same columns as in the cache CSV
    with open(dms_cache_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        ws.append(reader.fieldnames)
        row_count = 0
        for row in reader:
            ws.append([row.get(col, "") for col in reader.fieldnames])
            row_count += 1
    if row_count == 0:
        raise ValueError("No rows found in DMSDocuments cache CSV.")
    wb.save(output_path)
    return str(output_path)
import csv
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent



"""
def resolve_existing_articles_file(base_dir, target):
    
    #Resolve existing articles destination file for target in {'prod','test'}.
    #Returns None when target is empty/none.
    
    normalized = str(target or "").strip().lower()
    if normalized in {"", "none", "off", "false", "0"}:
        return None

    mapping = {
        "prod": "existing_articles_PROD.csv",
        "test": "existing_articles_TEST.csv",
    }
    if normalized not in mapping:
        raise ValueError("existing_articles_target must be one of: none, prod, test")

    cache_dir = Path(base_dir) / "data" / "processed" / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / mapping[normalized]
"""

def _read_existing_article_keys(path: Path):
    """
    Read existing article keys (artnr values) from existing articles CSV.
    Handles both old format (single column 'artnr') and new format (columns: 'artnr,zeichnr').
    Returns set of artnr keys.
    """
    keys = set()
    if not path.exists():
        return keys

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=",")
        rows = list(reader)

    if not rows:
        return keys

    header = [str(c or "").strip().lower() for c in rows[0]]
    data_rows = rows[1:] if header and "artnr" in header else rows
    idx = header.index("artnr") if header and "artnr" in header else 0

    for row in data_rows:
        if idx < len(row):
            key = str(row[idx] or "").strip()
            if key:
                keys.add(key)
    return keys



def _detect_encoding(file_path: Path):
    """Detect encoding for uploaded text/csv file (UTF-8, UTF-16, Latin1 fallback)."""
    with open(file_path, "rb") as f:
        raw = f.read(4096)
    # BOM detection
    if raw.startswith(b"\xff\xfe"):
        return "utf-16"
    if raw.startswith(b"\xfe\xff"):
        return "utf-16"
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except Exception:
        pass
    try:
        raw.decode("latin1")
        return "latin1"
    except Exception:
        pass
    return "utf-8"  # fallback


def _extract_artnr_from_excel(file_path: Path):
    """
    Extract article numbers from the first worksheet of an Excel file.
    Supports files with or without a header row.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if not wb.worksheets:
            return set()

        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows(values_only=True):
            normalized = [str(c or "").strip() for c in row]
            if any(normalized):
                rows.append(normalized)

        if not rows:
            return set()

        candidate_headers = {"artnr", "artikelnummer", "artikel_nr", "artikel-nr", "article", "article_no"}
        first_row_lower = [c.lower() for c in rows[0]]

        has_header = any(h in candidate_headers for h in first_row_lower)
        idx = 0
        if has_header:
            for i, h in enumerate(first_row_lower):
                if h in candidate_headers:
                    idx = i
                    break

        data_rows = rows[1:] if has_header else rows
        result = set()
        for row in data_rows:
            if idx < len(row):
                artnr = str(row[idx] or "").strip()
                if artnr:
                    result.add(artnr)
        return result
    finally:
        wb.close()


def _extract_artnr_zeichnr_from_excel(file_path: Path):
    """
    Extract (artnr, zeichnr) tuples from the first worksheet of an Excel file.
    Supports files with or without a header row.
    Returns list of tuples: [(artnr, zeichnr), ...]
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if not wb.worksheets:
            return []

        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows(values_only=True):
            normalized = [str(c or "").strip() for c in row]
            if any(normalized):
                rows.append(normalized)

        if not rows:
            return []

        candidate_headers_artnr = {"artnr", "artikelnummer", "artikel_nr", "artikel-nr", "article", "article_no"}
        candidate_headers_zeichnr = {"zeichnr", "zeichnung", "zeichnungsnummer", "drawing", "drawing_no"}
        first_row_lower = [c.lower() for c in rows[0]]

        has_header = any(h in candidate_headers_artnr for h in first_row_lower)
        artnr_idx = 0
        zeichnr_idx = -1

        if has_header:
            for i, h in enumerate(first_row_lower):
                if h in candidate_headers_artnr:
                    artnr_idx = i
                elif h in candidate_headers_zeichnr:
                    zeichnr_idx = i

        data_rows = rows[1:] if has_header else rows
        result = []
        for row in data_rows:
            if artnr_idx < len(row):
                artnr = str(row[artnr_idx] or "").strip()
                if artnr:
                    zeichnr = ""
                    if zeichnr_idx >= 0 and zeichnr_idx < len(row):
                        zeichnr = str(row[zeichnr_idx] or "").strip()
                    result.append((artnr, zeichnr))
        return result
    finally:
        wb.close()

def _extract_artnr_from_file(file_path: Path):
    """
    Extract article numbers from uploaded iFAS artikelstamm files.
    - Text/CSV: supports ; , tab and pipe delimiters, with or without headers.
    - Excel: supports .xlsx/.xlsm/.xltx/.xltm (first worksheet).
    Handles UTF-8, UTF-16, and Latin1 encodings for text files.
    """
    if str(file_path.suffix or "").lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _extract_artnr_from_excel(file_path)

    encoding = _detect_encoding(file_path)
    with file_path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        delimiter = ";"
        scores = {
            ";": sample.count(";"),
            ",": sample.count(","),
            "\t": sample.count("\t"),
            "|": sample.count("|"),
        }
        delimiter = max(scores, key=scores.get) if sample else ";"

        reader = csv.reader(f, delimiter=delimiter)
        rows = [row for row in reader if any(str(c or "").strip() for c in row)]

    if not rows:
        return set()

    header = [str(c or "").strip().lower() for c in rows[0]]
    candidate_headers = {"artnr", "artikelnummer", "artikel_nr", "artikel-nr", "article", "article_no"}
    has_header = any(h in candidate_headers for h in header)
    idx = 0
    if has_header:
        for i, h in enumerate(header):
            if h in candidate_headers:
                idx = i
                break

    data_rows = rows[1:] if has_header else rows
    result = set()
    for row in data_rows:
        if idx < len(row):
            artnr = str(row[idx] or "").strip()
            if artnr:
                result.add(artnr)
    return result


def _extract_artnr_zeichnr_from_file(file_path: Path):
    """
    Extract (artnr, zeichnr) tuples from uploaded iFAS artikelstamm files.
    - Text/CSV: supports ; , tab and pipe delimiters, with or without headers.
    - Excel: supports .xlsx/.xlsm/.xltx/.xltm (first worksheet).
    Returns list of tuples: [(artnr, zeichnr), ...]
    """
    if str(file_path.suffix or "").lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _extract_artnr_zeichnr_from_excel(file_path)

    encoding = _detect_encoding(file_path)
    with file_path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        delimiter = ";"
        scores = {
            ";": sample.count(";"),
            ",": sample.count(","),
            "\t": sample.count("\t"),
            "|": sample.count("|"),
        }
        delimiter = max(scores, key=scores.get) if sample else ";"

        reader = csv.reader(f, delimiter=delimiter)
        rows = [row for row in reader if any(str(c or "").strip() for c in row)]

    if not rows:
        return []

    header = [str(c or "").strip().lower() for c in rows[0]]
    candidate_headers_artnr = {"artnr", "artikelnummer", "artikel_nr", "artikel-nr", "article", "article_no"}
    candidate_headers_zeichnr = {"zeichnr", "zeichnung", "zeichnungsnummer", "drawing", "drawing_no"}
    
    has_header = any(h in candidate_headers_artnr for h in header)
    artnr_idx = 0
    zeichnr_idx = -1

    if has_header:
        for i, h in enumerate(header):
            if h in candidate_headers_artnr:
                artnr_idx = i
            elif h in candidate_headers_zeichnr:
                zeichnr_idx = i

    data_rows = rows[1:] if has_header else rows
    result = []
    for row in data_rows:
        if artnr_idx < len(row):
            artnr = str(row[artnr_idx] or "").strip()
            if artnr:
                zeichnr = ""
                if zeichnr_idx >= 0 and zeichnr_idx < len(row):
                    zeichnr = str(row[zeichnr_idx] or "").strip()
                result.append((artnr, zeichnr))
    return result


def append_article_list_to_existing(article_list_path, existing_articles_path):
    """
    Add articles from article_list.csv to existing articles csv (comma-separated, columns: artnr,zeichnr).
    Reads artnr and zeichnr from article_list.csv (semicolon-delimited).
    Returns number of new rows appended.
    """
    article_list_path = Path(article_list_path)
    existing_articles_path = Path(existing_articles_path)
    existing_articles_path.parent.mkdir(parents=True, exist_ok=True)

    existing_keys = _read_existing_article_keys(existing_articles_path)

    article_data = []
    if article_list_path.exists():
        with article_list_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                artnr = str((row or {}).get("artnr", "") or "").strip()
                zeichnr = str((row or {}).get("zeichnr", "") or "").strip()
                if artnr:
                    article_data.append((artnr, zeichnr))

    to_add = [(artnr, zeichnr) for artnr, zeichnr in article_data if artnr not in existing_keys]

    if not existing_articles_path.exists() or existing_articles_path.stat().st_size == 0:
        with existing_articles_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=",")
            writer.writerow(["artnr", "zeichnr"])

    if to_add:
        with existing_articles_path.open("a", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=",")
            for artnr, zeichnr in to_add:
                writer.writerow([artnr, zeichnr])

    return len(to_add)


def update_existing_articles_from_ifas_upload(upload_path, existing_articles_path, article_list_path=None):
    """
    Overwrite existing_articles_{PROD|TEST}.csv with artnr and zeichnr from the uploaded file.
    Format: comma-separated with columns: artnr,zeichnr
    Returns summary dict.
    """
    upload_path = Path(upload_path)
    existing_articles_path = Path(existing_articles_path)

    uploaded_data = _extract_artnr_zeichnr_from_file(upload_path)
    uploaded_data = sorted(uploaded_data, key=lambda x: x[0])  # Sort by artnr

    existing_articles_path.parent.mkdir(parents=True, exist_ok=True)
    with existing_articles_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=",")
        writer.writerow(["artnr", "zeichnr"])
        for artnr, zeichnr in uploaded_data:
            writer.writerow([artnr, zeichnr])

    return {
        "uploaded_count": len(uploaded_data),
        "added_count": len(uploaded_data),
        "existing_total": len(uploaded_data),
        "target_file": str(existing_articles_path),
    }


def _sanitize_name(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(value or ""))
    cleaned = cleaned.strip("_")
    return cleaned or "module"


def _normalize_key(value: str) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _resolve_cache_csv_path(cache_dir: Path, sheet_name: str) -> Path:
    exact_path = cache_dir / f"{sheet_name}_cache.csv"
    if exact_path.exists():
        return exact_path

    target_key = _normalize_key(sheet_name)
    for candidate in cache_dir.glob("*_cache.csv"):
        candidate_sheet = candidate.stem[:-6] if candidate.stem.endswith("_cache") else candidate.stem
        if _normalize_key(candidate_sheet) == target_key:
            return candidate

    return exact_path


def _read_cache_rows(cache_csv_path: Path):
    if not cache_csv_path.exists():
        return [], []
    with cache_csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    return rows, headers


def _build_referenz_lookup(template_wb):
    lookup = {}
    if "Referenz" not in template_wb.sheetnames:
        return lookup

    ws = template_wb["Referenz"]
    # Expected structure: col A = worksheet name, col B = db table name.
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        sheet_name = str(row[0] or "").strip()
        db_name = str(row[1] or "").strip()
        if not sheet_name:
            continue
        lookup[_normalize_key(sheet_name)] = db_name
    return lookup


def _detect_header_row(first_7_rows, cache_headers):
    # Pick the row inside 1..7 with the highest overlap with cache CSV headers.
    cache_set = {_normalize_key(h) for h in cache_headers if str(h or "").strip()}
    best_row = 7
    best_score = -1

    for row_idx, row_vals in first_7_rows.items():
        template_headers = {_normalize_key(v) for v in row_vals if str(v or "").strip()}
        score = len(template_headers.intersection(cache_set))
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row



def _copy_sheet_layout(template_ws, out_ws, first_rows_to_copy=7, safe_mode=True):
    # Copy only cell values from the first rows — no styles, no protection.
    # This avoids openpyxl IndexError caused by protection/style ID mismatches.
    max_col = template_ws.max_column
    for r in range(1, first_rows_to_copy + 1):
        for c in range(1, max_col + 1):
            src = template_ws.cell(row=r, column=c)
            out_ws.cell(row=r, column=c, value=src.value)

    return max_col


def create_import_excel_from_templates(
    template_path,
    cache_dir,
    active_sheet_names,
    output_path,
    default_column_width=14,
    safe_mode=True,
):
    """
    Build iFAS import workbook from template and cache files.
    - Copies first 7 rows and layout exactly from template sheets.
    - Matches cache columns to template headers found in first 7 rows.
    - Leaves unmatched template columns empty below row 7.
    - Keeps/updates hidden Referenz sheet with active sheet -> db table rows.
    - Forces all data cells to text format to preserve leading zeroes.
    """
    template_path = Path(template_path)
    cache_dir = Path(cache_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_template = openpyxl.load_workbook(template_path)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # Build referenz_lookup as a list of (sheet_name, db_table) from the template Referenz sheet
    referenz_lookup = []
    if "Referenz" in wb_template.sheetnames:
        ws_ref_template = wb_template["Referenz"]
        for r in range(2, ws_ref_template.max_row + 1):
            sheet = str(ws_ref_template.cell(row=r, column=1).value or "").strip()
            db = str(ws_ref_template.cell(row=r, column=2).value or "").strip()
            if sheet:
                referenz_lookup.append((sheet, db))

    # Build a dict for fast lookup
    referenz_dict = {_normalize_key(sheet): db for sheet, db in referenz_lookup}

    # Process active sheets and collect their db table names
    active_referenz_rows = []

    print("[DEBUG] Active sheet names:", active_sheet_names)
    for sheet_name in active_sheet_names:
        if sheet_name not in wb_template.sheetnames:
            print(f"[DEBUG] Sheet '{sheet_name}' not in template")
            continue

        cache_csv_path = _resolve_cache_csv_path(cache_dir, sheet_name)
        cache_rows, cache_headers = _read_cache_rows(cache_csv_path)
        print(f"[DEBUG] Sheet: {sheet_name}")
        print(f"  [DEBUG] cache_rows: {len(cache_rows)}")
        print(f"  [DEBUG] cache_headers: {cache_headers}")
        if not cache_rows:
            print(f"  [ERROR] No entries in article list for module. Module already exists in iFAS {{TEST/PROD}} or no new articles to export for sheet '{sheet_name}'.")
            raise ValueError(f"No entries in article list for module. Module already exists in iFAS {{TEST/PROD}} or no new articles to export for sheet '{sheet_name}'.")

        ws_template = wb_template[sheet_name]
        ws_out = wb_out.create_sheet(sheet_name)

        max_col = _copy_sheet_layout(ws_template, ws_out, first_rows_to_copy=7, safe_mode=safe_mode)
        print(f"  [DEBUG] max_col={max_col}")

        first_7 = {
            r: [ws_template.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            for r in range(1, 8)
        }
        header_row_idx = _detect_header_row(first_7, cache_headers)
        print(f"  [DEBUG] Detected header_row_idx={header_row_idx}")

        template_header_cells = [ws_template.cell(row=header_row_idx, column=c).value for c in range(1, max_col + 1)]
        print(f"  [DEBUG] template_header_cells: {template_header_cells}")

        # Explicit mapping for headers that differ between cache and template
        header_aliases = {
            "zusatztextart 1": "Zusatztextart 1_x000d_\n(ISZTAId)",
            "zusatztextart 2": "Zusatztextart 2_x000d_\n(ISZTAId)",
            "sind zusatztexte rtf text?": "Sind Zusatztexte RTF Text? _x000d_\n(0 = nein, 1 = ja)",
        }
        cache_header_map = {}
        for h in cache_headers:
            norm_h = _normalize_key(h)
            # Use alias if present, else original
            if h in header_aliases:
                cache_header_map[_normalize_key(header_aliases[h])] = h
            else:
                cache_header_map[norm_h] = h

        data_start_row = 8
        # Map each template header to a cache header (log if not found)
        unmapped_headers = []
        header_to_cache = []
        for tpl_header in template_header_cells:
            tpl_key = _normalize_key(tpl_header)
            if tpl_key:
                cache_header = cache_header_map.get(tpl_key)
                if cache_header is None:
                    unmapped_headers.append(tpl_header)
                header_to_cache.append(cache_header)
            else:
                unmapped_headers.append(tpl_header)
                header_to_cache.append(None)

        print(f"  [DEBUG] header_to_cache mapping: {header_to_cache}")
        if unmapped_headers:
            print(f"  [WARNING] Unmapped template headers in sheet '{sheet_name}': {unmapped_headers}")

        for row_offset, cache_row in enumerate(cache_rows):
            excel_row = data_start_row + row_offset
            for col_idx, cache_header in enumerate(header_to_cache, start=1):
                value = cache_row.get(cache_header, "") if cache_header else ""
                try:
                    cell = ws_out.cell(row=excel_row, column=col_idx, value="" if value is None else str(value))
                    cell.number_format = "@"
                except Exception as e:
                    print(f"[DEBUG] Error setting cell({excel_row}, {col_idx}): {e}")

        for col_idx, tpl_header in enumerate(template_header_cells, start=1):
            tpl_key = _normalize_key(tpl_header)
            if not tpl_key:
                continue
            try:
                letter = get_column_letter(col_idx)
                for row in range(1, ws_out.max_row + 1):
                    cell = ws_out.cell(row=row, column=col_idx)
                    cell.number_format = "@"
            except Exception as e:
                print(f"[DEBUG] Error setting format for column {col_idx}: {e}")

        for col_idx, tpl_header in enumerate(template_header_cells, start=1):
            tpl_key = _normalize_key(tpl_header)
            if not tpl_key:
                continue
            letter = get_column_letter(col_idx)
            if ws_out.column_dimensions[letter].width is None:
                ws_out.column_dimensions[letter].width = default_column_width
            ws_out.column_dimensions[letter].width = max(ws_out.column_dimensions[letter].width or 0, default_column_width)

        # Add to referenz rows for output
        db_table = referenz_dict.get(_normalize_key(sheet_name), sheet_name)
        active_referenz_rows.append((sheet_name, db_table))

    # Create Referenz sheet from scratch: header + only active sheets
    ws_ref_out = wb_out.create_sheet("Referenz")
    ws_ref_out.cell(row=1, column=1, value="Worksheet").number_format = "@"
    ws_ref_out.cell(row=1, column=2, value="Tabelle").number_format = "@"
    for idx, (sheet, db) in enumerate(active_referenz_rows, start=2):
        ws_ref_out.cell(row=idx, column=1, value=sheet).number_format = "@"
        ws_ref_out.cell(row=idx, column=2, value=db).number_format = "@"

    try:
        print(f"[DEBUG] Saving workbook with {len(wb_out.sheetnames)} sheets: {wb_out.sheetnames}")
        for sheet_name in wb_out.sheetnames:
            ws = wb_out[sheet_name]
            print(f"[DEBUG] Sheet '{sheet_name}': max_row={ws.max_row}, max_col={ws.max_column}")
        wb_out.save(output_path)
        print(f"[DEBUG] Successfully saved Excel to {output_path}")
    except Exception as e:
        print(f"[ERROR] openpyxl failed to save Excel: {e}")
        import traceback
        traceback.print_exc()
        raise
    return output_path


def archive_module_export(module_artnr, excel_path, partlist_csv_path, partlist_tree_path, archive_dir):
    """
    Save artifacts under data/archive and create a zip named with the module.
    Returns: (archive_folder_path, zip_path)
    """
    module_name = _sanitize_name(module_artnr)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    archive_dir = Path(archive_dir)
    archive_dir.mkdir(parents=True, exist_ok=True)

    # Instead of creating a folder, just zip the files directly and remove any temp folders
    excel_path = Path(excel_path)
    partlist_csv_path = Path(partlist_csv_path)
    partlist_tree_path = Path(partlist_tree_path)

    zip_path = archive_dir / f"{module_name}_{stamp}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Add Excel
        if excel_path.exists():
            zf.write(excel_path, arcname=f"{module_name}_iFAS_import.xlsx")
        # Add partlist CSV
        if partlist_csv_path.exists():
            zf.write(partlist_csv_path, arcname=f"partlist_{module_name}.csv")
        # Add partlist tree
        if partlist_tree_path.exists():
            zf.write(partlist_tree_path, arcname=f"partlist_tree_{module_name}.txt")

    # Optionally, remove any old archive folders for this module (cleanup)
    # for folder in archive_dir.glob(f"{module_name}_*"):
    #     if folder.is_dir():
    #         shutil.rmtree(folder, ignore_errors=True)

    return None, zip_path


def _normalize_sheet_name_key(value: str) -> str:
    import unicodedata

    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(ch for ch in normalized if ch.isalnum()).lower()


def _is_bom_sheet_name(sheet_name: str) -> bool:
    key = _normalize_sheet_name_key(sheet_name)
    return "stuckliste" in key or "stueckliste" in key


def _read_active_sheet_names(active_sheets_path: Path):
    if not active_sheets_path.exists():
        return []
    with active_sheets_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    if not rows:
        return []
    return [str(v or "").strip() for v in rows[0] if str(v or "").strip()]


def _find_cache_csv_for_sheet(cache_dir: Path, sheet_name: str):
    # Prefer the exact expected naming first.
    exact = cache_dir / f"{sheet_name}_cache.csv"
    if exact.exists():
        return exact

    # Fallback: normalized matching across existing *_cache.csv files.
    target_key = _normalize_sheet_name_key(sheet_name)
    for path in cache_dir.glob("*_cache.csv"):
        stem = path.stem.replace("_cache", "")
        if _normalize_sheet_name_key(stem) == target_key:
            return path
    return None


def create_partlist_import_excel_from_cache(
    template_xlsx_path: Path,
    cache_dir: Path,
    active_sheets_path: Path,
    output_xlsx_path: Path,
    first_rows_to_copy: int = 7,
):
    """
    Build the partlist import workbook from BOM cache sheets.

    Rules:
    - include only active BOM sheets (Stueckliste/Stuecklisten*) plus mandatory Referenz.
    - copy first template rows (default 1..7) for each active BOM worksheet.
    - map data columns by matching cache header names to row-7 header values in template.
    - rebuild Referenz with header row: Worksheet;Tabelle and active-sheet rows.
    """
    template_xlsx_path = Path(template_xlsx_path)
    cache_dir = Path(cache_dir)
    active_sheets_path = Path(active_sheets_path)
    output_xlsx_path = Path(output_xlsx_path)
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if not template_xlsx_path.exists():
        raise FileNotFoundError(f"Template not found: {template_xlsx_path}")

    wb_template = openpyxl.load_workbook(template_xlsx_path)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    active_sheet_names = _read_active_sheet_names(active_sheets_path)
    active_bom_sheets = [s for s in active_sheet_names if _is_bom_sheet_name(s)]

    # Keep order and remove duplicates.
    seen = set()
    ordered_bom_sheets = []
    for sheet in active_bom_sheets:
        k = _normalize_sheet_name_key(sheet)
        if k in seen:
            continue
        seen.add(k)
        ordered_bom_sheets.append(sheet)

    # Map template referenz table names by worksheet.
    referenz_lookup = _build_referenz_lookup(wb_template)
    referenz_rows = []

    for sheet_name in ordered_bom_sheets:
        # Resolve worksheet name exactly from template sheet names by normalized key.
        template_sheet = None
        target_key = _normalize_sheet_name_key(sheet_name)
        for candidate in wb_template.sheetnames:
            if _normalize_sheet_name_key(candidate) == target_key:
                template_sheet = candidate
                break
        if not template_sheet:
            continue

        cache_csv_path = _find_cache_csv_for_sheet(cache_dir, template_sheet)
        cache_rows, cache_headers = _read_cache_rows(cache_csv_path) if cache_csv_path else ([], [])

        ws_template = wb_template[template_sheet]
        ws_out = wb_out.create_sheet(template_sheet)
        max_col = _copy_sheet_layout(ws_template, ws_out, first_rows_to_copy=first_rows_to_copy, safe_mode=True)

        # Header row is defined by requirement as row 7.
        template_headers_row7 = [ws_template.cell(row=7, column=c).value for c in range(1, max_col + 1)]
        cache_header_map = {_normalize_sheet_name_key(h): h for h in cache_headers}

        for row_offset, cache_row in enumerate(cache_rows):
            excel_row = first_rows_to_copy + 1 + row_offset
            for col_idx, template_header in enumerate(template_headers_row7, start=1):
                t_key = _normalize_sheet_name_key(template_header)
                cache_header = cache_header_map.get(t_key)
                value = cache_row.get(cache_header, "") if cache_header else ""
                cell = ws_out.cell(row=excel_row, column=col_idx, value="" if value is None else str(value))
                cell.number_format = "@"

        db_table = referenz_lookup.get(_normalize_key(template_sheet), template_sheet)
        referenz_rows.append((template_sheet, db_table))

    # Mandatory Referenz sheet.
    ws_ref = wb_out.create_sheet("Referenz")
    ws_ref.cell(row=1, column=1, value="Worksheet").number_format = "@"
    ws_ref.cell(row=1, column=2, value="Tabelle").number_format = "@"
    for idx, (sheet, db_table) in enumerate(referenz_rows, start=2):
        ws_ref.cell(row=idx, column=1, value=sheet).number_format = "@"
        ws_ref.cell(row=idx, column=2, value=db_table).number_format = "@"

    wb_out.save(output_xlsx_path)
    return output_xlsx_path

def create_partlist_excel_from_template(
    partlist_csv_path: Path,
    template_xlsx_path: Path,
    output_xlsx_path: Path,
    first_rows_to_copy: int = 7
):
    """
    Generate an Excel file from partlist.csv using the template layout.
    Preserves the first template rows, clears any stale sample data below them,
    and fills the import columns expected by the template.
    """
    import openpyxl
    from copy import copy

    wb = openpyxl.load_workbook(template_xlsx_path)
    ws_out = wb.active

    data_start_row = first_rows_to_copy + 1
    max_col = ws_out.max_column

    # Capture the template's first data row as the style/source baseline.
    template_row_values = [ws_out.cell(row=data_start_row, column=c).value for c in range(1, max_col + 1)]
    template_row_styles = [copy(ws_out.cell(row=data_start_row, column=c)._style) for c in range(1, max_col + 1)]

    # Remove stale sample data from the template before writing new rows.
    for row_index in range(data_start_row, ws_out.max_row + 1):
        for column_index in range(1, max_col + 1):
            ws_out.cell(row=row_index, column=column_index).value = None

    # Read partlist.csv
    with open(partlist_csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        part_rows = list(reader)

    def _normalize_header(value):
        import unicodedata
        normalized = unicodedata.normalize("NFKD", str(value or ""))
        return "".join(ch for ch in normalized if ch.isalnum()).lower()

    header_to_value = {
        _normalize_header("Stücklistennummer"): lambda row: row.get("stulinr", "") or "",
        _normalize_header("Artikelnummer"): lambda row: row.get("artnr", "") or "",
        _normalize_header("Bezeichnung [de]"): lambda row: row.get("artbez1", "") or "",
        _normalize_header("Menge"): lambda row: row.get("menge", "") or "",
        _normalize_header("SL-Bezugsmenge"): lambda row: row.get("menge", "") or "",
        _normalize_header("Stücklistenstatus"): lambda row: "10",
        _normalize_header("Org.Einheit Identifier"): lambda row: "JOS",
    }

    # Write data into the template's existing structure.
    for idx, row in enumerate(part_rows):
        excel_row = data_start_row + idx
        for column_index in range(1, max_col + 1):
            cell = ws_out.cell(row=excel_row, column=column_index)
            if column_index - 1 < len(template_row_styles):
                cell._style = copy(template_row_styles[column_index - 1])

            template_header = template_row_values[column_index - 1] if column_index - 1 < len(template_row_values) else ""
            normalized_header = _normalize_header(template_header)
            value_getter = header_to_value.get(normalized_header)
            if value_getter is not None:
                cell.value = value_getter(row)
            else:
                cell.value = template_header if False else cell.value

        # Ensure the known template columns are set even if the template header text changes slightly.
        ws_out.cell(row=excel_row, column=2).value = row.get("stulinr", "") or ""
        ws_out.cell(row=excel_row, column=4).value = row.get("artnr", "") or ""
        ws_out.cell(row=excel_row, column=5).value = "10"
        ws_out.cell(row=excel_row, column=6).value = row.get("artbez1", "") or ""
        ws_out.cell(row=excel_row, column=11).value = row.get("menge", "") or ""
        ws_out.cell(row=excel_row, column=12).value = "JOS"

    wb.save(output_xlsx_path)
    return output_xlsx_path
