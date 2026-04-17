import openpyxl
from pathlib import Path
import csv

def extract_worksheets(xlsx_path, out_csv_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    with open(out_csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['sheet_name'])
        for sheet in wb.sheetnames:
            writer.writerow([sheet])
    print(f"Extracted {len(wb.sheetnames)} sheets to {out_csv_path}")

if __name__ == "__main__":
    xlsx_path = Path("data/raw/templates/Partlist_import_template_App.xlsx")
    out_csv_path = Path("config/sheets.csv")
    extract_worksheets(xlsx_path, out_csv_path)
