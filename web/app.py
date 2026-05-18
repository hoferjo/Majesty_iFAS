###imports
from statistics import mode
import yaml
from etl.create import ArticleCreator
from etl.extract import auto_update_dbf_csv
from etl.extract import extract_dbf_headers_and_rows, save_folder_tree
from etl.validate import (
    load_article_group_validation_queue,
    save_article_group_item,
    add_group_entry,
    overwrite_bezeichnungen_from_config,
    load_article_group_validation_queue_from_csv,
    save_article_group_item_to_csv,
)
import threading
from pathlib import Path
from fastapi import Body
import sys
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import shutil
import os
import json
import csv
import time
import unicodedata
from fastapi import Query
from fastapi.responses import JSONResponse



### Path setup, 
BASE_DIR = Path(__file__).parent.parent

def load_yaml(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _parse_bool_flag(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _read_bezeichnungen_anpassen_setting() -> bool:
    try:
        current_settings = load_yaml(settings_path) or {}
        features = current_settings.get("features") or {}
        return _parse_bool_flag(features.get("bezeichnungen_anpassen"), default=False)
    except Exception:
        return False


def _write_bezeichnungen_anpassen_setting(enabled: bool) -> None:
    current_settings = load_yaml(settings_path) or {}
    features = current_settings.get("features") or {}
    features["bezeichnungen_anpassen"] = bool(enabled)
    current_settings["features"] = features
    with open(settings_path, "w", encoding="utf-8", newline="") as handle:
        yaml.safe_dump(current_settings, handle, allow_unicode=True, sort_keys=False)
    
settings_path = BASE_DIR / "config" / "settings.yaml"
settings = load_yaml(settings_path)
if not settings:
    raise RuntimeError(f"Failed to load settings from {settings_path}. File is empty or invalid.")
try:
    raw_dir = settings["paths"]["raw_dir"]
    artikelstamm_path = settings["files"]["artikelstamm"]
except Exception as e:
    raise RuntimeError(f"Error reading keys from settings.yaml: {e}")
etl_dir = BASE_DIR / settings["paths"]["etl_dir"]
sys.path.append(str(etl_dir))
 

def get_sheet_cache_csv(sheet: str, blocked: bool = False):
    if not blocked:
        return sheets_output_dir / f'{sheet}_cache.csv'
    else:
        return sheets_output_dir / f'{sheet}_cache_blocked.csv'
    


def get_path(file: str = None, mode: str = None, base: str = BASE_DIR):
    
    normalized_mode = str(mode or "").strip().lower()
    if normalized_mode.endswith("_mode"):
        normalized_mode = normalized_mode[:-5]

    file_key = file
    folder_key = f"{file}_dir"

    if file == "article list":
        folder_key = "article_list_dir"
        file_key = f"article_list_{normalized_mode}_mode"

    if file == "blocked articles":
        folder_key = "article_list_dir"
        file_key = f"blocked_articles_{normalized_mode}_mode"

    if file == "existing articles":
        if normalized_mode in {"", "none"}:
            return None
        folder_key = "existing_articles_dir"
        file_key = f"existing_articles_{normalized_mode}"

    if file == "partlist":
        folder_key = "partlist_dir"
        file_key = f"partlist_{normalized_mode}_mode"

    if file == "partlisttree":
        folder_key = "partlist_tree_dir"
        file_key = f"partlist_{normalized_mode}_mode_tree"

    if file in {"artikelstamm_import_template", "partlist_import_template"}:
        folder_key = "template_dir"
        file_key = file

    if file_key and folder_key:
        return base / settings["paths"][folder_key] / settings["files"][file_key]

    return os.error(f"Invalid file key: {file}. No matching path configuration found in settings.yaml.")


      

#raw paths

artikelstamm_path = get_path("artikelstamm")
stuecklistenstamm_path = get_path("stuecklistenstamm")
waren_artikelgruppe_path = get_path("waren_artikelgruppe")
lieferanten_mapping_path = get_path("lieferanten_mapping")

#processed paths
cache_dir_path = BASE_DIR / settings["paths"]["cache_dir"]
sheets_path = BASE_DIR / "config" / "sheets.csv"
active_sheets_path = BASE_DIR / "config" / "active_sheets.csv"
sheets_output_dir = BASE_DIR / settings["paths"]["sheets_dir"]

#template paths
template_dir = BASE_DIR / settings["paths"]["template_dir"]
artikelstamm_import_template_path = get_path("artikelstamm_import_template")
partlist_import_template_path = get_path("partlist_import_template")

#output_paths
output_dir = BASE_DIR / settings["paths"]["output_dir"]
artikelstamm_output_path = output_dir / settings["files"]["artikelstamm_output"]
partlist_output_path = output_dir / settings["files"]["partlist_output"]

#article list paths
"""
article_list_article_path = get_path(BASE_DIR, "article list", "article")
article_list_module_path = get_path(BASE_DIR, "article list", "module")
article_list_creation_path = get_path(BASE_DIR, "article list", "creation")
"""
#blocked articles paths
""" 
blocked_articles_article_path = get_path(BASE_DIR, "blocked articles", "article")
blocked_articles_module_path = get_path(BASE_DIR, "blocked articles", "module")
blocked_articles_creation_path = get_path(BASE_DIR, "blocked articles", "creation")
"""

#existing articles paths
"""
existing_articles_PROD_path = get_path(BASE_DIR, "existing articles", "PROD")
existing_articles_TEST_path = get_path(BASE_DIR, "existing articles", "TEST")
"""

#partlist paths
"""
partlist_article_mode_path = get_path("partlist", "article", BASE_DIR)
partlist_module_mode_path = get_path("partlist", "module", BASE_DIR)
partlist_creation_mode_path = get_path("partlist", "creation", BASE_DIR)
"""

#partlisttree paths
"""
partlist_article_mode_tree_path = get_path("partlisttree", "article", BASE_DIR)
partlist_module_mode_tree_path = get_path("partlisttree", "module", BASE_DIR)
partlist_creation_mode_tree_path = get_path("partlisttree", "creation", BASE_DIR)
"""



#output paths

from etl.transform import process_module_structure, build_sheet_cache_CSV, build_bom_sheet_cache, build_docs_download_cache_csv
from etl.load import (
    create_import_excel_from_templates,
    archive_module_export,
    append_article_list_to_existing,
    update_existing_articles_from_ifas_upload,
    create_partlist_excel_from_template,
    create_partlist_import_excel_from_cache,
)
from etl.validate import load_article_validation_queue, save_article_validation_item
from etl.create import create_from_drawingtree, add_unique_artnr





app = FastAPI()


def _load_article_creator() -> ArticleCreator:
    config_dir = Path(__file__).parent.parent / "config"
    return ArticleCreator(config_dir)


def _get_effective_template(creator: ArticleCreator, type_name: str, group_name: str | None = None):
    tpl = creator.get_template(type_name)
    if tpl and group_name:
        tpl = creator.merge_defaults(group_name, tpl)
    return tpl



# --- Root Article/Module Creation Workflow ---
from fastapi import APIRouter
from typing import Dict, Any

@app.post("/api/create-root-article")
def api_create_root_article():
    """
    Returns the required fields for all active sheets, with default values if available.
    """
    BASE_DIR = Path(__file__).parent.parent
    _, headers, aliases, flags, active_headers, _ = _load_sheets_config(BASE_DIR)
    # For each active sheet, collect its fields (headers)
    active_sheets = []
    for idx, header in enumerate(headers):
        if header not in active_headers:
            continue
        alias = str(aliases[idx]).strip() if idx < len(aliases) else ""
        sheet_name = _resolve_sheet_for_mapping(BASE_DIR, header, alias)
        if not sheet_name:
            continue
        # Try to load default values from config/create_defaults.yaml if present
        defaults = {}
        defaults_path = BASE_DIR/ "config" / "create_defaults.yaml"
        if defaults_path.exists():
            import yaml
            with open(defaults_path, "r", encoding="utf-8") as f:
                all_defaults = yaml.safe_load(f) or {}
            defaults = all_defaults.get(sheet_name, {})
        # For now, just use the header names as fields
        active_sheets.append({
            "sheet": sheet_name,
            "fields": list(defaults.keys()) if defaults else [],
            "defaults": defaults,
        })
    return {"status": "ok", "sheets": active_sheets}


@app.post("/api/create-from-tree")
def api_create_from_tree(data: dict = Body({})):
    """Parse a drawing tree and export a CSV of drawings.
    Optional JSON body: { 
        "tree_file": "path/to/tree.txt", 
        "output_csv": "path/to/out.csv",
        "folder": "folder_name_filter" (optional, to filter by category/subcategory)
    }
    If omitted, uses default tree and output locations under the project.
    """
    base = Path(__file__).parent.parent
    tree_file = data.get("tree_file")
    output_csv = data.get("output_csv")
    folder_filter = data.get("folder")  # Optional folder filter

    default_tree = base / "data" / "raw" / "drawings_tree" / "tree_Eigenprodukte_Jost_Artikel.txt"
    default_out = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree.csv"

    tree_path = Path(tree_file) if tree_file else default_tree
    out_path = Path(output_csv) if output_csv else default_out

    try:
        create_from_drawingtree(tree_path, out_path, folder_filter=folder_filter)
        # Count rows written
        written = 0
        if out_path.exists():
            with out_path.open("r", encoding="utf-8-sig") as f:
                # subtract header
                written = sum(1 for _ in f) - 1
                if written < 0:
                    written = 0
        filter_msg = f" from folder '{folder_filter}'" if folder_filter else ""
        return {"status": "success", "message": f"Created {written} drawing rows{filter_msg}.", "output": str(out_path)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/api/add-unique-artnr")
def api_add_unique_artnr(data: dict = Body({})):
    """Create `article_list_from_tree_unique.csv` by adding artnr where possible.

    JSON body options:
      - existing_mode: 'PROD' or 'TEST' (default PROD)
      - tree_csv: optional path to input tree CSV
      - out_csv: optional path for unique output CSV
    """
    base = Path(__file__).parent.parent
    existing_mode = (data.get("existing_mode") or "PROD").strip().upper()
    tree_csv = data.get("tree_csv")
    out_csv = data.get("out_csv")

    default_tree = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree.csv"
    default_out = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"

    tree_path = Path(tree_csv) if tree_csv else default_tree
    out_path = Path(out_csv) if out_csv else default_out

    try:
        count = add_unique_artnr(tree_path, out_path, existing_mode=existing_mode)
        return {"status": "success", "message": f"Wrote {count} unique drawings to {out_path}", "output": str(out_path)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/api/save-tree")
def api_save_tree(data: dict = Body({})):
    """Save a fresh tree from a filesystem folder to the project's raw/drawings_tree file.

    Optional JSON body keys:
        - root_folder: path to walk (defaults to T:\\01 Jost AG\\05 Produktion\\11 Eigenprodukte Jost)
        - out_file: target file path (defaults to data/raw/drawings_tree/tree_Eigenprodukte_Jost_Artikel.txt)
    """
    base = Path(__file__).parent.parent
    root_folder = data.get("root_folder")
    out_file = data.get("out_file")

    default_root = r"T:\\01 Jost AG\\05 Produktion\\11 Eigenprodukte Jost"
    default_out = base / "data" / "raw" / "drawings_tree" / "tree_Eigenprodukte_Jost_Artikel.txt"

    root = Path(root_folder) if root_folder else Path(default_root)
    outp = Path(out_file) if out_file else default_out

    try:
        save_folder_tree(root, outp)
        return {"status": "success", "message": f"Wrote tree for {root} to {outp}", "output": str(outp)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/api/validate/groups/from-file")
def api_validate_groups_from_file(data: dict = Body({})):
    """Load group validation queue from an input CSV (created articles).

    JSON body options:
      - input_file: path to input CSV (optional)
    If omitted, uses default `data/processed/cache/created_articles/article_list_from_tree_unique.csv`.
    """
    base = Path(__file__).parent.parent
    input_file = data.get("input_file")
    default_in = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"
    csv_path = Path(input_file) if input_file else default_in

    try:
        result = load_article_group_validation_queue_from_csv(base, csv_path)
        return result
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/api/validate/groups/from-file/save")
def api_validate_groups_from_file_save(data: dict = Body({})):
    """Save group validation updates into a file specific for created-articles validation.

    JSON body:
      - artnr: article number
      - updates: dict with hauptgruppe/untergruppe/spezifikation/bezeichnungselemente
      - out_file: optional path to overrides CSV to write (defaults to created_articles dir)
    """
    base = Path(__file__).parent.parent
    artnr = data.get("artnr")
    updates = data.get("updates") or {
        "hauptgruppe": data.get("hauptgruppe", ""),
        "untergruppe": data.get("untergruppe", ""),
        "spezifikation": data.get("spezifikation", ""),
        "bezeichnungselemente": data.get("bezeichnungselemente", []),
    }
    # write directly into the input CSV by default (or into provided input_file)
    input_file = data.get("input_file")
    default_in = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"
    out_path = Path(input_file) if input_file else default_in

    try:
        result = save_article_group_item_to_csv(base, artnr, updates, out_path)
        return result
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/api/build-sheet-cache-creation")
def api_build_sheet_cache_creation(data: dict = Body({})):
    """Build sheet cache for created articles (creation mode).

    Reads article_list_from_tree_unique.csv and builds cache CSVs for all active sheets.
    JSON body options:
      - input_file: path to article list CSV (optional, defaults to created_articles unique CSV)
    """
    base = Path(__file__).parent.parent
    input_file = data.get("input_file")
    active_sheets_file = data.get("active_sheets_file")

    default_in = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"
    default_active_sheets = base / "config" / "active_sheets.csv"

    csv_path = Path(input_file) if input_file else default_in
    active_sheets_path = Path(active_sheets_file) if active_sheets_file else default_active_sheets

    try:
        # Load active sheets
        if not active_sheets_path.exists():
            return JSONResponse(status_code=400, content={"status": "error", "message": f"Active sheets file not found: {active_sheets_path}"})

        with open(active_sheets_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            active_sheets_list = next(reader, []) or []

        if not active_sheets_list:
            return JSONResponse(status_code=400, content={"status": "error", "message": "No active sheets configured"})

        # Check article CSV exists
        if not csv_path.exists():
            return JSONResponse(status_code=400, content={"status": "error", "message": f"Article list not found: {csv_path}"})

        # Build sheet cache
        build_sheet_cache_CSV(str(csv_path), active_sheets_list)

        # Count rows in each sheet cache
        sheets_dir = base / "data" / "processed" / "cache" / "sheets"
        sheet_counts = []
        for sheet in active_sheets_list:
            cache_path = sheets_dir / f"{sheet}_cache.csv"
            count = 0
            if cache_path.exists():
                with open(cache_path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f, delimiter=";")
                    count = sum(1 for _ in reader)
            sheet_counts.append({"sheet": sheet, "rows": count})

        return {
            "status": "success",
            "message": f"Built cache for {len(active_sheets_list)} sheets from {csv_path.name}",
            "sheets": sheet_counts,
            "cache_dir": str(sheets_dir),
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.get("/download-creation-excel")
def download_creation_excel():
    """Download import Excel for created articles (creation mode).
    
    Uses article_list_from_tree_unique.csv and active sheets to build the import workbook.
    """
    base = Path(__file__).parent.parent
    template_path = base / "data" / "raw" / "templates" / "Vorlage_edit_jhofer.xlsx"
    cache_dir = base / "data" / "processed" / "cache" / "sheets"
    temp_export_dir = base / "data" / "processed" / "cache" / "export"
    temp_export_dir.mkdir(parents=True, exist_ok=True)
    creation_list_path = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"

    if not template_path.exists():
        return JSONResponse(status_code=404, content={"status": "error", "message": f"Template not found: {template_path}"})

    if not cache_dir.exists():
        return JSONResponse(status_code=400, content={"status": "error", "message": "Sheet cache not found. Build sheet cache first."})

    try:
        active_sheets_list = _resolved_active_sheet_names(base)

        if not active_sheets_list:
            return JSONResponse(status_code=400, content={"status": "error", "message": "No active sheets configured"})

        if creation_list_path.exists():
            build_sheet_cache_CSV(str(creation_list_path), active_sheets_list)

        tag = "created_articles"
        output_xlsx = temp_export_dir / f"{tag}_iFAS_import.xlsx"
        
        create_import_excel_from_templates(
            template_path=template_path,
            cache_dir=cache_dir,
            active_sheet_names=active_sheets_list,
            output_path=output_xlsx,
            default_column_width=14,
        )

        return FileResponse(
            str(output_xlsx),
            filename=f"{tag}_iFAS_import.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.get("/download-article-data-xlsx")
def download_article_data_xlsx(folder_name: str = Query("")):
    """Download article data from cache as xlsx. Exports all Article-category cache sheets and
    injects Zeichnungsnummer, Zeichnungsindex and Bezeichnung 1 from the created-articles list
    when those columns are missing.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import re

    base = Path(__file__).parent.parent
    cache_dir = base / "data" / "processed" / "cache" / "sheets"
    export_dir = base / "data" / "processed" / "cache" / "export"
    export_dir.mkdir(parents=True, exist_ok=True)
    template_path = base / "data" / "raw" / "templates" / "Vorlage_edit_jhofer.xlsx"

    safe_suffix = str(folder_name or "").strip().replace("\\", "/").strip("/")
    safe_suffix = Path(safe_suffix).name if safe_suffix else ""
    safe_suffix = re.sub(r"[^0-9A-Za-zÄÖÜäöüß._-]+", "_", safe_suffix).strip("_")
    filename = f"Import_Artikelstamm_{safe_suffix}.xlsx" if safe_suffix else "Import_Artikelstamm.xlsx"

    # Load creation article list to enrich sheet rows
    creation_list_path = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"
    article_map = {}
    if creation_list_path.exists():
        with open(creation_list_path, encoding="utf-8-sig") as f:
            rdr = csv.DictReader(f, delimiter=';')
            for r in rdr:
                artnr = (r.get('artnr') or r.get('Artikelnummer') or '').strip()
                if artnr:
                    article_map[artnr] = r

    try:
        if creation_list_path.exists() and template_path.exists():
            active_sheets_list = _resolved_active_sheet_names(base)
            if active_sheets_list:
                build_sheet_cache_CSV(str(creation_list_path), active_sheets_list)
                output_xlsx = export_dir / filename
                create_import_excel_from_templates(
                    template_path=template_path,
                    cache_dir=cache_dir,
                    active_sheet_names=active_sheets_list,
                    output_path=output_xlsx,
                    default_column_width=14,
                )

                return FileResponse(
                    str(output_xlsx),
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                        "Pragma": "no-cache",
                        "Expires": "0",
                    },
                )

        wb = openpyxl.Workbook()
        first = True

        # Determine article sheet names from config mappings
        article_mappings_dir = base / 'config' / 'sheet_mappings' / 'Article'
        article_sheet_names = []
        if article_mappings_dir.exists():
            for p in article_mappings_dir.glob('mapping_plan_*.csv'):
                name = p.stem.replace('mapping_plan_', '')
                article_sheet_names.append(name)

        # Fallback: include common article sheets if no mapping files found
        if not article_sheet_names:
            article_sheet_names = ['Artikelstamm', 'ArtikelDisposteuerung', 'ArtikelLieferantendaten']

        for sheet_name in article_sheet_names:
            cache_file = cache_dir / f"{sheet_name}_cache.csv"
            if not cache_file.exists():
                continue

            with open(cache_file, encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f, delimiter=';')
                headers = reader.fieldnames or []

                # Ensure Zeichnungsnummer / Zeichnungsindex / Bezeichnung 1 [de] are present
                extra_map = {
                    'Zeichnungsnummer': 'zeichnr',
                    'Zeichnungsindex': 'zeichindex',
                    'Bezeichnung 1 [de]': 'artbez1',
                }
                for ex in extra_map:
                    if ex not in headers:
                        headers.append(ex)

                # Create or select worksheet
                if first:
                    ws = wb.active
                    ws.title = sheet_name
                    first = False
                else:
                    ws = wb.create_sheet(title=sheet_name)

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Write rows
                for row_idx, row in enumerate(reader, 2):
                    artnr_val = (row.get('Artikelnummer') or row.get('artnr') or '').strip()
                    for col_idx, header in enumerate(headers, 1):
                        if header in extra_map:
                            mapped_key = extra_map[header]
                            value = ''
                            if artnr_val and artnr_val in article_map:
                                value = article_map[artnr_val].get(mapped_key, '')
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))

                # Adjust column widths for this sheet
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

        output_xlsx = export_dir / filename
        wb.save(output_xlsx)

        return FileResponse(
            str(output_xlsx),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.get("/download-document-data-xlsx")
def download_document_data_xlsx(folder_name: str = Query("")):
    """Download document data from cache as xlsx."""
    import re
    
    base = Path(__file__).parent.parent
    cache_dir = base / "data" / "processed" / "cache" / "sheets"
    export_dir = base / "data" / "processed" / "cache" / "export"
    export_dir.mkdir(parents=True, exist_ok=True)
    creation_list_path = base / "data" / "processed" / "cache" / "created_articles" / "article_list_from_tree_unique.csv"
    template_xlsx_path = base / "data" / "raw" / "templates" / "Template_DMS_Dokumentenverknüpfung_V1.xlsx"

    safe_suffix = str(folder_name or "").strip().replace("\\", "/").strip("/")
    safe_suffix = Path(safe_suffix).name if safe_suffix else ""
    safe_suffix = re.sub(r"[^0-9A-Za-zÄÖÜäöüß._-]+", "_", safe_suffix).strip("_")
    filename = f"Import_Document_{safe_suffix}.xlsx" if safe_suffix else "Import_Document.xlsx"
    
    try:
        if creation_list_path.exists() and template_xlsx_path.exists():
            docs_cache_path = cache_dir / "DMSDocuments_cache.csv"
            docs_cache_result = build_docs_download_cache_csv(
                article_list_path=creation_list_path,
                output_csv_path=docs_cache_path,
            )
            if docs_cache_result.get("rows_written", 0) > 0:
                output_xlsx = export_dir / filename
                export_docs_xlsx(
                    [],
                    output_path=output_xlsx,
                    template_xlsx_path=template_xlsx_path,
                    cache_csv_path=docs_cache_path,
                )
                return FileResponse(
                    str(output_xlsx),
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                        "Pragma": "no-cache",
                        "Expires": "0",
                    },
                )

        cache_file = cache_dir / "DMSDocuments_cache.csv"
        if not cache_file.exists():
            cache_file = cache_dir / "DMSDocuments.csv"

        if not cache_file.exists():
            return JSONResponse(status_code=400, content={"status": "error", "message": "Document cache not found"})

        # Legacy flat workbook fallback.
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documents"

        with open(cache_file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            headers = reader.fieldnames or []

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_idx, row in enumerate(reader, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

        output_xlsx = export_dir / "document_data.xlsx"
        wb.save(output_xlsx)
        
        return FileResponse(
            str(output_xlsx),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})

# Global flag to clear article cache files after import
ARTICLE_CACHE_CLEAR = 0

# Keeps the latest generated module article rows per artnr in this process.
_MODULE_ARTICLES_CACHE = {}
_MODULE_EXISTING_TARGET_CACHE = {}
_SEARCH_CACHE = {"path": None, "mtime": None, "rows": []}


def _resolve_partlist_template_path():
    # Prefer configured template, then known fallback names found in template_dir.
    candidates = [
        partlist_import_template_path,
        template_dir / "Partlist_import_template_App.xlsx",
        template_dir / "VorlageStücklisteV1.xlsx",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return Path(candidate)
    return partlist_import_template_path


def _load_search_rows(csv_path: Path):
    try:
        mtime = csv_path.stat().st_mtime
    except OSError:
        return []

    if (
        _SEARCH_CACHE["path"] == str(csv_path)
        and _SEARCH_CACHE["mtime"] == mtime
        and _SEARCH_CACHE["rows"]
    ):
        return _SEARCH_CACHE["rows"]

    rows = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            rows.append({
                "artnr": row.get("artnr", "") or "",
                "artbez1": row.get("artbez1", "") or "",
                "artbez2": row.get("artbez2", "") or "",
                "artbez3": row.get("artbez3", "") or "",
                "zeichnr": row.get("zeichnr", "") or "",
            })

    _SEARCH_CACHE["path"] = str(csv_path)
    _SEARCH_CACHE["mtime"] = mtime
    _SEARCH_CACHE["rows"] = rows
    return rows


def _strip_trailing_empty(values):
    trimmed = list(values)
    while trimmed and str(trimmed[-1]).strip() == "":
        trimmed.pop()
    return trimmed


def _load_sheets_config(base: Path):
    if not sheets_path.exists():
        raise FileNotFoundError(f"sheets.csv not found at {sheets_path}")

    with open(sheets_path, "r", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f, delimiter=';'))

    if not rows:
        raise ValueError("sheets.csv is empty.")

    headers = _strip_trailing_empty(rows[0])
    aliases = _strip_trailing_empty(rows[1]) if len(rows) > 1 else []
    flags = _strip_trailing_empty(rows[2]) if len(rows) > 2 else []

    active_headers = []
    for idx, header in enumerate(headers):
        flag = str(flags[idx]).strip().lower() if idx < len(flags) else "0"
        if flag in {"1", "true", "yes", "y", "on"}:
            active_headers.append(header)

    return sheets_path, headers, aliases, flags, active_headers, rows


def _mapping_exists(base: Path, sheet_name: str):
    return _resolve_mapping_sheet_name(base, sheet_name) is not None


def _resolve_mapping_sheet_name(base: Path, sheet_name: str):
    candidate = str(sheet_name or "").strip()
    if not candidate:
        return None

    direct_paths = [
        base / "config" / f"mapping_plan_{candidate}.csv",
        base / "config" / "sheet_mappings" / "Article" / f"mapping_plan_{candidate}.csv",
        base / "config" / "sheet_mappings" / "BOM" / f"mapping_plan_{candidate}.csv",
        base / "config" / "sheet_mappings" / "Workplan" / f"mapping_plan_{candidate}.csv",
    ]
    for p in direct_paths:
        if p.exists():
            return p.stem.replace("mapping_plan_", "")

    target_norm = _normalize_sheet_header(candidate)
    search_dirs = [
        base / "config",
        base / "config" / "sheet_mappings" / "Article",
        base / "config" / "sheet_mappings" / "BOM",
        base / "config" / "sheet_mappings" / "Workplan",
    ]
    for d in search_dirs:
        if not d.exists():
            continue
        for path in d.glob("mapping_plan_*.csv"):
            sheet = path.stem.replace("mapping_plan_", "")
            if _normalize_sheet_header(sheet) == target_norm:
                return sheet
    return None


def _resolve_sheet_for_mapping(base: Path, header: str, alias: str):
    candidates = []
    if alias:
        candidates.append(alias)
    if header:
        candidates.append(header)
        candidates.append(header.replace("_", ""))
    if alias and alias.endswith("steuerung"):
        candidates.append(alias + "en")
    if header and header.endswith("_steuerung"):
        candidates.append(header.replace("_", "") + "en")

    seen = set()
    for candidate in candidates:
        candidate = str(candidate).strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        resolved = _resolve_mapping_sheet_name(base, candidate)
        if resolved:
            return resolved
    return None


def _read_article_list_rows(article_list_path: Path):
    with open(article_list_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        return list(reader)


def _load_artikelstamm_details_map(artikelstamm_path: Path):
    details = {}
    if not artikelstamm_path.exists():
        return details
    with open(artikelstamm_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            artnr = str(row.get("artnr", "")).strip()
            if not artnr:
                continue
            details[artnr] = {
                "artnr": artnr,
                "artbez1": row.get("artbez1", "") or "",
                "artbez2": row.get("artbez2", "") or "",
                "artbez3": row.get("artbez3", "") or "",
                "artbezmem": row.get("artbezmem", "") or "",
                "zeichnr": row.get("zeichnr", "") or "",
            }
    return details


def _read_blocked_article_details(blocked_articles_path: Path, artikelstamm_path: Path):
    if not blocked_articles_path.exists():
        return []

    details_map = _load_artikelstamm_details_map(artikelstamm_path)
    blocked_items = []
    seen = set()

    with open(blocked_articles_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            artnr = str(row.get("artnr", "")).strip()
            if not artnr or artnr in seen:
                continue
            seen.add(artnr)

            enriched = details_map.get(artnr, {
                "artnr": artnr,
                "artbez1": row.get("artbez1", "") or "",
                "artbez2": "",
                "artbez3": "",
                "artbezmem": "",
                "zeichnr": row.get("zeichnr", "") or "",
            })
            blocked_items.append(enriched)

    return blocked_items


def _resolved_active_sheet_names(base: Path):
    _, headers, aliases, _, active_headers, _ = _load_sheets_config(base)
    active_set = set(active_headers)
    resolved = []
    for idx, header in enumerate(headers):
        if header not in active_set:
            continue
        alias = str(aliases[idx]).strip() if idx < len(aliases) else ""
        sheet_name = _resolve_sheet_for_mapping(base, header, alias)
        if sheet_name:
            resolved.append(sheet_name)
    return list(dict.fromkeys(resolved))


def _normalize_mode(mode: str | None, default: str = "module"):
    mode_value = str(mode or default).strip().lower()
    if mode_value == "article":
        return "article"
    if mode_value == "creation":
        return "creation"
    return "module"


def _cache_paths(base: Path, mode: str):
    normalized_mode = _normalize_mode(mode)

    return {
        "mode": normalized_mode,
        "article_list": get_path("article list", normalized_mode, base),
        "blocked_articles": get_path("blocked articles", normalized_mode, base),
        "partlist": get_path("partlist", normalized_mode, base),
        "partlist_tree": get_path("partlisttree", normalized_mode, base),
    }


def _build_bom_caches_if_available(base: Path, partlist_path: Path, article_list_path: Path | None):
    if not partlist_path or not Path(partlist_path).exists():
        return None

    article_list_csv = None
    if article_list_path and Path(article_list_path).exists():
        article_list_csv = str(article_list_path)

    return build_bom_sheet_cache(
        str(partlist_path),
        article_list_csv,
        base,
    )


def _reset_mode_cache_files(base: Path, mode: str):
    paths = _cache_paths(base, mode)
    headers = {
        "article_list": "artnr;artbez1;zeichnr\n",
        "blocked_articles": "artnr;artbez1;zeichnr\n",
        "partlist": "stulinr;posnr;menge;artnr;artbez1\n",
    }
    for key, header in headers.items():
        with open(paths[key], "w", encoding="utf-8") as f:
            f.write(header)
    if paths["partlist_tree"].exists():
        paths["partlist_tree"].unlink()
    return paths


def _count_csv_rows(path: Path):
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            return max(sum(1 for _ in f) - 1, 0)
    except Exception:
        return 0


def _normalize_existing_target(value: str):
    target = str(value or "none").strip().lower()
    if target not in {"none", "prod", "test"}:
        raise ValueError("existing_articles_target must be one of: none, prod, test")
    return target


def _normalize_sheet_header(value: str):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(ch for ch in normalized if ch.isalnum()).lower()


def _resolve_selected_header(raw_header: str, header_lookup: dict):
    key = _normalize_sheet_header(raw_header)
    if key in header_lookup:
        return header_lookup[key]

    aliases = {
        "stucklistenvarianten": "stuecklistevarianten",
        "stuecklistenvarianten": "stuecklistevarianten",
        "stucklstenauswahlvarianten": "stuecklisteauswahlvarianten",
        "stuecklstenauswahlvarianten": "stuecklisteauswahlvarianten",
        "stucklistenauswahlvarianten": "stuecklisteauswahlvarianten",
        "stuecklistenauswahlvarianten": "stuecklisteauswahlvarianten",
    }
    mapped = aliases.get(key, key)
    return header_lookup.get(mapped)

# Get the absolute path to the web directory
web_dir = Path(__file__).parent.absolute()

# Mount static files (JS, CSS)
app.mount("/static", StaticFiles(directory=str(web_dir / "static")), name="static")


@app.get("/", response_class=HTMLResponse)
def index():
    template_path = web_dir / "templates" / "index.html"
    return open(template_path, "r", encoding="utf-8").read()

@app.get("/settings", response_class=HTMLResponse)
def settings_page():
    template_path = web_dir / "templates" / "settings.html"
    return open(template_path, "r", encoding="utf-8").read()


@app.get("/upload", response_class=HTMLResponse)
def upload_page():
    template_path = web_dir / "templates" / "upload.html"
    return open(template_path, "r", encoding="utf-8").read()

@app.post("/upload-file")
def upload_file(file: UploadFile = File(...)):
    upload_dir = "data/uploaded_files"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"filename": file.filename, "status": "uploaded"}



@app.get("/search")
def search(query: str = Query(..., min_length=1), mode: str = Query("article")):
    # Path to the artikelstamm CSV (adjust as needed)
    csv_path = Path(__file__).parent.parent / "data" / "raw" / "artikelstamm" / "artikelstamm_majesty_2026_03_30.csv"
    results = []
    try:
        q = query.lower()
        for row in _load_search_rows(csv_path):
            if (
                q in row["artnr"].lower()
                or q in row["zeichnr"].lower()
                or q in row["artbez1"].lower()
                or q in row["artbez2"].lower()
                or q in row["artbez3"].lower()
            ):
                results.append({
                    "artnr": row["artnr"],
                    "artbez1": row["artbez1"],
                    "zeichnr": row["zeichnr"],
                })
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
    return {"results": results}


@app.get("/sheets-config")
def sheets_config():
    base = Path(__file__).parent.parent
    try:
        _, headers, _, _, active_headers, _ = _load_sheets_config(base)
        return {
            "status": "success",
            "headers": headers,
            "active_headers": active_headers,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.get("/api/bezeichnungen-anpassen")
def api_get_bezeichnungen_anpassen():
    return {
        "status": "success",
        "enabled": _read_bezeichnungen_anpassen_setting(),
    }


@app.post("/api/bezeichnungen-anpassen")
def api_set_bezeichnungen_anpassen(data: dict = Body(...)):
    enabled = _parse_bool_flag((data or {}).get("enabled"), default=False)
    try:
        _write_bezeichnungen_anpassen_setting(enabled)
        return {"status": "success", "enabled": enabled}
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})



# --- Generate Module endpoint ---
@app.post("/generate-module")
def generate_module(data: dict = Body(...)):

    artnr = data.get("artnr")
    existing_articles_target = _normalize_existing_target(data.get("existing_articles_target", "none"))
    replacement_map = data.get("replacement_map") or {}
    if not isinstance(replacement_map, dict):
        return JSONResponse(status_code=400, content={"status": "error", "message": "replacement_map must be an object"})

    normalized_replacement_map = {}
    for key, value in replacement_map.items():
        key_s = str(key or "").strip()
        value_s = str(value or "").strip()
        if key_s and value_s:
            normalized_replacement_map[key_s] = value_s

    if not artnr:
        return JSONResponse(status_code=400, content={"status": "error", "message": "No artnr provided"})



    ####
    cache_paths = _cache_paths(BASE_DIR, "module")
    ####

    partlist_path = get_path("partlist", "module")
    article_list_path = get_path("article list", "module")
    blocked_articles_path = get_path("blocked articles", "module")
    try:
        existing_articles_path = None
        if existing_articles_target != "none":
            existing_articles_path = get_path("existing articles", existing_articles_target, BASE_DIR)
        _MODULE_EXISTING_TARGET_CACHE[str(artnr)] = existing_articles_target

        process_module_structure(
            str(artnr),
            str(artikelstamm_path),
            str(stuecklistenstamm_path),
            str(article_list_path),
            str(partlist_path),
            mode = "module",
            existing_articles_path = str(existing_articles_path) if existing_articles_path else None,
            replacement_map = normalized_replacement_map,
            reset_files = True,
        )
        _MODULE_ARTICLES_CACHE[str(artnr)] = _read_article_list_rows(article_list_path)

        bom_result = _build_bom_caches_if_available(base=Path(__file__).parent.parent, partlist_path=partlist_path, article_list_path=article_list_path)

        article_count = _count_csv_rows(article_list_path)
        partlist_count = _count_csv_rows(partlist_path)
        blocked_count = _count_csv_rows(blocked_articles_path)
        blocked_articles_data = ""
        if blocked_count > 0 and blocked_articles_path.exists():
            with open(blocked_articles_path, "r", encoding="utf-8-sig") as f:
                blocked_articles_data = f.read()
        blocked_items = _read_blocked_article_details(blocked_articles_path, artikelstamm_path)

        msg = f"Modulestructure generated for {artnr} articles to migrate: {article_count} partlist entries: {partlist_count} blocked articles: {blocked_count}"
        if bom_result and isinstance(bom_result, dict) and not bom_result.get("error"):
            msg += f" | BOM sheets: {bom_result.get('stuecklisten_count', 0)} stücklisten, {bom_result.get('versionen_count', 0)} versionen, {bom_result.get('positionen_count', 0)} positionen"
        elif bom_result and isinstance(bom_result, dict) and bom_result.get("error"):
            msg += f" | BOM sheet error: {bom_result['error']}"
        return JSONResponse(status_code=200, content={
            "status": "success",
            "message": msg,
            "blocked_articles": blocked_articles_data,
            "blocked_items": blocked_items
        })
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/generate-module-apply-replacements")
def generate_module_apply_replacements(data: dict = Body(...)):
    artnr = data.get("artnr")
    existing_articles_target = _normalize_existing_target(data.get("existing_articles_target", "none"))
    replacement_map = data.get("replacement_map") or {}

    if not artnr:
        return JSONResponse(status_code=400, content={"status": "error", "message": "No artnr provided"})
    if not isinstance(replacement_map, dict):
        return JSONResponse(status_code=400, content={"status": "error", "message": "replacement_map must be an object"})

    normalized_replacement_map = {}
    for key, value in replacement_map.items():
        key_s = str(key or "").strip()
        # Accept dict for textartikel action
        if isinstance(value, dict) and value.get("textartikel"):
            normalized_replacement_map[key_s] = {"textartikel": True}
        else:
            value_s = str(value or "").strip()
            if key_s and value_s:
                normalized_replacement_map[key_s] = value_s

    cache_paths = _cache_paths(BASE_DIR, "module")
    partlist_path = cache_paths["partlist"]
    blocked_articles_path = cache_paths["blocked_articles"]
    article_list_path = cache_paths["article_list"]

    try:
        existing_articles_path = None
        if existing_articles_target != "none":
            existing_articles_path = get_path("existing articles", existing_articles_target, BASE_DIR)
        _MODULE_EXISTING_TARGET_CACHE[str(artnr)] = existing_articles_target

        process_module_structure(
            str(artnr),
            str(artikelstamm_path),
            str(stuecklistenstamm_path),
            str(article_list_path),
            str(partlist_path),
            mode = "module",
            existing_articles_path = str(existing_articles_path) if existing_articles_path else None,
            replacement_map = normalized_replacement_map,
            reset_files = True,
        )

        _MODULE_ARTICLES_CACHE[str(artnr)] = _read_article_list_rows(article_list_path)

        bom_result = _build_bom_caches_if_available(base=Path(__file__).parent.parent, partlist_path=partlist_path, article_list_path=article_list_path)

        article_count = _count_csv_rows(article_list_path)
        partlist_count = _count_csv_rows(partlist_path)
        blocked_count = _count_csv_rows(blocked_articles_path)
        blocked_articles_data = ""
        if blocked_count > 0 and blocked_articles_path.exists():
            with open(blocked_articles_path, "r", encoding="utf-8-sig") as f:
                blocked_articles_data = f.read()

        blocked_items = _read_blocked_article_details(blocked_articles_path, artikelstamm_path)

        msg = (
            f"Modulestructure generated for {artnr}: "
            f"articles to migrate: {article_count}, partlist entries: {partlist_count}, blocked articles: {blocked_count}"
        )
        if bom_result and isinstance(bom_result, dict) and not bom_result.get("error"):
            msg += f" | BOM sheets: {bom_result.get('stuecklisten_count', 0)} stücklisten, {bom_result.get('versionen_count', 0)} versionen, {bom_result.get('positionen_count', 0)} positionen"
        elif bom_result and isinstance(bom_result, dict) and bom_result.get("error"):
            msg += f" | BOM sheet error: {bom_result['error']}"

        return JSONResponse(status_code=200, content={
            "status": "success",
            "message": msg,
            "blocked_articles": blocked_articles_data,
            "blocked_items": blocked_items,
        })
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/upload-ifas-artikelstamm")
def upload_ifas_artikelstamm(
    file: UploadFile = File(...),
    target_env: str = Form("test"),
):
    env = str(target_env or "").strip().lower()
    if env not in {"prod", "test"}:
        return JSONResponse(status_code=400, content={"status": "error", "message": "target_env must be prod or test"})

    try:
        target_file = get_path("existing articles", env, BASE_DIR)
        if not target_file:
            return JSONResponse(status_code=400, content={"status": "error", "message": "No target file resolved"})

        upload_dir = BASE_DIR / "data" / "uploaded_files"
        upload_dir.mkdir(parents=True, exist_ok=True)
        uploaded_path = upload_dir / file.filename

        with open(uploaded_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        summary = update_existing_articles_from_ifas_upload(
            upload_path=uploaded_path,
            existing_articles_path=target_file,
        
        )

        return {
            "status": "success",
            "message": (
                f"Updated existing articles {env.upper()}: "
                f"uploaded={summary['uploaded_count']}, "
                f"added={summary['added_count']}"
            ),
            "summary": summary,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@app.post("/generate-module-data")
def generate_module_data(data: dict = Body(...)):
    artnr = data.get("artnr")
    selected_headers = data.get("selected_headers", [])
    mode = _normalize_mode(data.get("mode"), "module")
    existing_articles_target = _normalize_existing_target(data.get("existing_articles_target", "none"))
    if not artnr:
        return {"status": "error", "message": "No artnr provided"}
    if not isinstance(selected_headers, list):
        return {"status": "error", "message": "selected_headers must be a list"}

    cache_paths = _cache_paths(BASE_DIR, mode)
    blocked_articles_path = cache_paths["blocked_articles"]
    article_list_path = cache_paths["article_list"]

    try:
        if not article_list_path.exists():
            return {
                "status": "error",
                "message": f"{article_list_path.name} not found. Generate {mode} structure first."
            }

        if not sheets_path.exists():
            return {
                "status": "error",
                "message": "sheets.csv not found in config."
            }

        sheets_output_dir.mkdir(parents=True, exist_ok=True)

        _, headers, aliases, _, _, rows = _load_sheets_config(BASE_DIR)
        header_lookup = {_normalize_sheet_header(h): h for h in headers}
        selected_set = set()
        unknown = []
        for raw in selected_headers:
            source = str(raw).strip()
            if not source:
                continue
            resolved = _resolve_selected_header(source, header_lookup)
            if resolved:
                selected_set.add(resolved)
            else:
                unknown.append(source)

        if unknown:
            return {
                "status": "error",
                "message": f"Unknown sheet headers selected: {', '.join(sorted(unknown))}"
            }

        # Ensure at least 3 rows and normalize row lengths.
        while len(rows) < 3:
            rows.append([])
        rows[0] = headers
        if len(rows[1]) < len(headers):
            rows[1] = rows[1] + [""] * (len(headers) - len(rows[1]))
        rows[1] = rows[1][:len(headers)]
        rows[2] = ["1" if header in selected_set else "0" for header in headers]

        with open(sheets_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=';')
            for row in rows:
                writer.writerow(row)

        # Renew active_sheets.csv with only marked column headers.
        selected_in_order = [h for h in headers if h in selected_set]
        with open(active_sheets_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(selected_in_order)

        if not selected_in_order:
            return {
                "status": "error",
                "message": "No sheets selected."
            }

        if mode == "article" and existing_articles_target != "none":
            _ = get_path("existing articles", existing_articles_target, BASE_DIR)

        # Build using resolvable mapping names from selected columns.
        build_sheet_names = []
        for idx, header in enumerate(headers):
            if header not in selected_set:
                continue
            alias = str(aliases[idx]).strip() if idx < len(aliases) else ""
            resolved = _resolve_sheet_for_mapping(BASE_DIR, header, alias)
            if resolved:
                build_sheet_names.append(resolved)

        build_sheet_names = list(dict.fromkeys(build_sheet_names))
        if not build_sheet_names:
            return {
                "status": "error",
                "message": "Selected sheets have no matching mapping_plan_*.csv files."
            }

        cached_articles = _MODULE_ARTICLES_CACHE.get(str(artnr))
        build_sheet_cache_CSV(str(article_list_path), build_sheet_names, articles=cached_articles)

        # DMS documents are generated from the creation article list separately so that
        # Bezeichnung and Dokumenten Speicherort are always populated from the drawing tree.
        if "DMSDocuments" in build_sheet_names:
            docs_cache_path = sheets_output_dir / "DMSDocuments_cache.csv"
            build_docs_download_cache_csv(
                article_list_path=article_list_path,
                output_csv_path=docs_cache_path,
                articles=cached_articles or _read_article_list_rows(article_list_path),
            )
        # Also generate for blocked articles
        if blocked_articles_path.exists():
            with blocked_articles_path.open("r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=';')
                blocked_articles = list(reader)
            if blocked_articles:
                build_sheet_cache_CSV(str(blocked_articles_path), build_sheet_names, articles=blocked_articles)
        # After generating, count entries in each sheet's cache CSV
        sheet_entry_counts = []
        for sheet in build_sheet_names:
            cache_csv = get_sheet_cache_csv(sheet)
            count = 0
            if cache_csv.exists():
                with cache_csv.open("r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.reader(f, delimiter=";")
                    next(reader, None)  # skip header
                    count = sum(1 for _ in reader)
            # Count blocked from *_cache_blocked.csv
            blocked_cache_csv = get_sheet_cache_csv(sheet, blocked = True)
            blocked_count = 0
            if blocked_cache_csv.exists():
                with blocked_cache_csv.open("r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.reader(f, delimiter=";")
                    next(reader, None)
                    blocked_count = sum(1 for _ in reader)
            sheet_entry_counts.append(f"{sheet}: {count} entries, blocked: {blocked_count}")
        # Generate BOM sheets only when BOM-related headers are selected.
        bom_result = None
        selected_norm = {_normalize_sheet_header(h) for h in selected_set}
        bom_norm_aliases = {
            "stuecklisten",
            "stueckliste",
            "stuecklistenvarianten",
            "stuecklistevarianten",
            "stuecklistenversionen",
            "stuecklistenverwendung",
            "stuecklstenauswahlvarianten",
            "stuecklistenauswahlvarianten",
            "stuecklisteauswahlvarianten",
            "stuecklistenpositionen",
        }
        should_build_bom = bool(selected_norm.intersection(bom_norm_aliases))
        if should_build_bom:
            try:
                bom_result = _build_bom_caches_if_available(
                    base=BASE_DIR,
                    partlist_path=cache_paths["partlist"],
                    article_list_path=cache_paths["article_list"],
                )
            except Exception as bom_e:
                bom_result = {"error": str(bom_e)}

        msg = "Module data generated. " + ", ".join(sheet_entry_counts)
        if bom_result and isinstance(bom_result, dict) and not bom_result.get("error"):
            msg += f" | BOM sheets: {bom_result.get('stuecklisten_count', 0)} stücklisten, {bom_result.get('versionen_count', 0)} versionen, {bom_result.get('positionen_count', 0)} positionen"
        elif bom_result and bom_result.get("error"):
            msg += f" | BOM sheet error: {bom_result['error']}"
        return {
            "status": "success",
            "message": msg
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/download-partlist-excel")
def download_partlist_excel(
    mode: str = Query("module"),
    artnr: str | None = Query(None),
):
    normalized_mode = _normalize_mode(mode, "module")
    template_path = _resolve_partlist_template_path()
    if not template_path.exists():
        return JSONResponse(status_code=404, content={"error": "partlist.csv or template not found."})

    temp_export_dir = BASE_DIR / "data" / "processed" / "csv" / "cache" / "export"
    temp_export_dir.mkdir(parents=True, exist_ok=True)
    tag = "".join(ch if ch.isalnum() or ch in {"-", "_", "."} else "_" for ch in str(artnr or normalized_mode))
    output_xlsx = temp_export_dir / f"partlist_export_{tag}.xlsx"
    try:
        create_partlist_import_excel_from_cache(
            template_xlsx_path=template_path,
            cache_dir=sheets_output_dir,
            active_sheets_path=active_sheets_path,
            output_xlsx_path=output_xlsx,
            first_rows_to_copy=7,
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

    return FileResponse(
        str(output_xlsx),
        filename=f"partlist_export_{tag}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/download-module-export")
def download_module_export(
    artnr: str = Query(..., min_length=1),
    existing_articles_target: str = Query("none"),
):
    requested_target = _normalize_existing_target(existing_articles_target)
    cached_target = _MODULE_EXISTING_TARGET_CACHE.get(str(artnr), "none")
    effective_target = requested_target if requested_target != "none" else cached_target
    print(f"[DEBUG] /download-module-export triggered for artnr={artnr}, target={effective_target}")
    base = Path(__file__).parent.parent

    template_path = base / "data" / "raw" / "templates" / "Vorlage_edit_jhofer.xlsx"
    cache_dir = sheets_output_dir
    cache_paths = _cache_paths(base, "module")
    partlist_path = cache_paths["partlist"]
    partlist_tree_path = cache_paths["partlist_tree"]
    archive_dir = base / "data" / "archive"
    temp_export_dir = base / "data" / "processed" / "csv" / "cache" / "export"
    temp_export_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.exists():
        return JSONResponse(status_code=404, content={"status": "error", "message": f"Template not found: {template_path}"})

    try:
        sheet_names = _resolved_active_sheet_names(base)
        if not sheet_names:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "No active sheets resolved. Generate module data first or check settings."},
            )

        temp_excel_path = temp_export_dir / f"{artnr}_iFAS_import.xlsx"
        create_import_excel_from_templates(
            template_path=template_path,
            cache_dir=cache_dir,
            active_sheet_names=sheet_names,
            output_path=temp_excel_path,
            default_column_width=14,
        )

        # Existing articles append now only happens in /download-module-excel

        _, zip_path = archive_module_export(
            module_artnr=artnr,
            excel_path=temp_excel_path,
            partlist_csv_path=partlist_path,
            partlist_tree_path=partlist_tree_path,
            archive_dir=archive_dir,
        )

        return FileResponse(
            str(zip_path),
            media_type="application/zip",
            filename=zip_path.name,
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})
    

@app.get("/download-module-excel")
def download_module_excel(
    artnr: str = Query(..., min_length=1),
    existing_articles_target: str = Query("none"),
    mode: str = Query("module"),
):
    mode = _normalize_mode(mode, "module")
    article_list_path = _cache_paths(BASE_DIR, mode)["article_list"]
    requested_target = _normalize_existing_target(existing_articles_target)
    cached_target = _MODULE_EXISTING_TARGET_CACHE.get(str(artnr), "none")
    effective_target = requested_target if requested_target != "none" else cached_target
    print(f"[DEBUG] /download-module-excel triggered for artnr={artnr}, target={effective_target}, mode={mode}")
    base = Path(__file__).parent.parent

    template_path = base / "data" / "raw" / "templates" / "Vorlage_edit_jhofer.xlsx"
    cache_dir = sheets_output_dir
    cache_paths = _cache_paths(base, mode)
    partlist_path = cache_paths["partlist"]
    partlist_tree_path = cache_paths["partlist_tree"]
    archive_dir = base / "data" / "archive"
    temp_export_dir = base / "data" / "processed" / "csv" / "cache" / "export"
    temp_export_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.exists():
        return JSONResponse(status_code=404, content={"status": "error", "message": f"Template not found: {template_path}"})

    global ARTICLE_CACHE_CLEAR
    try:
        sheet_names = _resolved_active_sheet_names(base)
        if not sheet_names:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "No active sheets resolved. Generate module data first or check settings."},
            )

        print(f"[DEBUG] /download-module-excel: artnr={artnr}, active_sheet_names={sheet_names}")
        temp_excel_path = temp_export_dir / f"{artnr}_iFAS_import.xlsx"
        print(f"[DEBUG] Creating Excel: template={template_path}, cache_dir={cache_dir}, output={temp_excel_path}")
        create_import_excel_from_templates(
            template_path=template_path,
            cache_dir=cache_dir,
            active_sheet_names=sheet_names,
            output_path=temp_excel_path,
            default_column_width=14,
        )
        print(f"[DEBUG] Excel creation complete for {artnr}")

        if mode == "module" and effective_target in {"prod", "test"}:
            target_file = get_path("existing articles", effective_target, base)
            append_article_list_to_existing(article_list_path, target_file)

        # Archive in background after sending Excel
        def archive_job():
            try:
                archive_module_export(
                    module_artnr=artnr,
                    excel_path=temp_excel_path,
                    partlist_csv_path=partlist_path,
                    partlist_tree_path=partlist_tree_path,
                    archive_dir=archive_dir,
                )
                print(f"[DEBUG] Archive complete for artnr={artnr}")
            except Exception as e:
                print(f"[ERROR] Archive failed for artnr={artnr}: {e}")

        threading.Thread(target=archive_job, daemon=True).start()

        if mode == "article":
            _reset_mode_cache_files(base, "article")
            ARTICLE_CACHE_CLEAR = 1

        return FileResponse(
            str(temp_excel_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=temp_excel_path.name,
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})
    
    
# --- Majesty DBF Extraction Endpoint ---
from etl.extract import auto_update_dbf_csv

@app.post("/update-majesty-data")
def update_majesty_data():
    try:
        updated = auto_update_dbf_csv()
        if updated:
            return {"status": "success", "message": f"Updated: {', '.join(updated)}"}
        else:
            return {"status": "success", "message": "No DBF files needed updating."}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    
    
@app.post("/hard-update-majesty-data")
def hard_update_majesty_data():
    try:
        updated = auto_update_dbf_csv(force=True)
        if updated:
            return {"status": "success", "message": f"Hard updated: {', '.join(updated)}"}
        else:
            # Check if there are any DBF files at all
            from pathlib import Path
            from etl.extract import dbfs_path
            dbf_dir = Path(dbfs_path)
            dbf_files = list(dbf_dir.glob("*.dbf"))
            if not dbf_files:
                return {"status": "error", "message": "No DBF files found in the Majesty DBF directory. Check your configuration and file locations."}
            else:
                return {"status": "success", "message": "No DBF files needed updating."}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    
@app.get("/download-partlist-tree")
def download_partlist_tree(artnr: str = Query(..., min_length=1), mode: str = Query("module")):
    base = Path(__file__).parent.parent
    partlist_tree_path = _cache_paths(base, mode)["partlist_tree"]
    # Optionally, support per-artnr tree files if needed:
    # partlist_tree_path = base / "data" / "processed" / "csv" / f"partlist_tree_{artnr}.txt"
    if not partlist_tree_path.exists():
        return JSONResponse(status_code=404, content={"error": f"partlist_tree.txt not found for {artnr} (mode={mode})"})
    try:
        return FileResponse(str(partlist_tree_path), filename=f"partlist_tree_{artnr}.txt", media_type="text/plain")
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
    

@app.post("/add-article")
def add_article(data: dict = Body(...)):
    artnr = data.get("artnr")
    if not artnr:
        return JSONResponse(status_code=400, content={"status": "error", "message": "No artnr provided"})
    requested_target = _normalize_existing_target(data.get("existing_articles_target", "none"))
    base= BASE_DIR
    article_list_path = get_path("article list", "article")
    artikelstamm_path = get_path("artikelstamm")
    cache_paths = _cache_paths(base, "article")
    article_list_path = cache_paths["article_list"]
    global ARTICLE_CACHE_CLEAR
    try:
        # If cache clear flag is set, clear article mode cache files and reset flag
        if ARTICLE_CACHE_CLEAR == 1:
            _reset_mode_cache_files(base, "article")
            ARTICLE_CACHE_CLEAR = 0
        import csv
        from etl.transform import check_utf8_file
        if not check_utf8_file(artikelstamm_path):
            return JSONResponse(status_code=500, content={"status": "error", "message": f"artikelstamm not valid UTF-8: {artikelstamm_path}"})
        # Read article-mode list entries to avoid duplicates
        existing_artnr = set()
        if article_list_path.exists():
            with open(article_list_path, 'r', encoding='utf-8-sig') as f:
                next(f, None)
                for line in f:
                    key = line.strip().split(';')[0].strip()
                    if key:
                        existing_artnr.add(key)
        # Check the selected active existing-articles target when provided.
        existing_targets_artnr = set()
        if requested_target != "none":
            target_file = get_path("existing articles", requested_target, base)
            if target_file and target_file.exists():
                with open(target_file, 'r', encoding='utf-8-sig') as f:
                    next(f, None)
                    for line in f:
                        key = line.strip().split(',')[0].strip()
                        if key:
                            existing_targets_artnr.add(key)
        # Find the article in artikelstamm
        found = False
        with open(artikelstamm_path, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                if str(row.get('artnr', '')).strip() == str(artnr).strip():
                    artbez1 = row.get('artbez1', '')
                    zeichnr = row.get('zeichnr', '')
                    if artnr in existing_artnr:
                        return {
                            "status": "success",
                            "message": f"Article {artnr} already exists in article_list_article_mode.csv. Nothing appended."
                        }
                    if artnr in existing_targets_artnr:
                        return {
                            "status": "success",
                            "message": f"Article {artnr} already exists in existing articles ({requested_target.upper()}). Nothing appended."
                        }
                    if artnr not in existing_artnr:
                        # Write header if file does not exist or is empty
                        write_header = not article_list_path.exists() or article_list_path.stat().st_size == 0
                        with open(article_list_path, 'a', encoding='utf-8') as out:
                            if write_header:
                                out.write('artnr;artbez1;zeichnr\n')
                            out.write(f"{artnr};{artbez1};{zeichnr}\n")
                    found = True
                    break
        if not found:
            return JSONResponse(status_code=404, content={"status": "error", "message": f"Article {artnr} not found in artikelstamm."})
        return {"status": "success", "message": f"Added article {artnr} to article_list.csv."}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})

# --- Article Mode: Generate Article Export and Partlist ---
@app.post("/generate-article")
def generate_article(data: dict = Body(...)):
    artnr_list = data.get("artnr_list")
    if not artnr_list or not isinstance(artnr_list, list):
        return JSONResponse(status_code=400, content={"status": "error", "message": "No article list provided"})

    base = Path(__file__).parent.parent
    existing_articles_target = _normalize_existing_target(data.get("existing_articles_target", "none"))
    cache_paths = _cache_paths(base, "article")
    article_list_path = cache_paths["article_list"]
    partlist_path = cache_paths["partlist"]
    partlist_tree_path = cache_paths["partlist_tree"]
    try:
        import csv
        from etl.transform import check_utf8_file, process_module_structure
        # Read existing articles to avoid duplicates
        existing_artnr = set()
        if article_list_path.exists():
            with open(article_list_path, 'r', encoding='utf-8-sig') as f:
                next(f, None)
                for line in f:
                    key = line.strip().split(';')[0].strip()
                    if key:
                        existing_artnr.add(key)
        existing_targets_artnr = set()
        if existing_articles_target != "none":
            target_file = get_path("existing articles", existing_articles_target, base)
            if target_file and target_file.exists():
                with open(target_file, 'r', encoding='utf-8-sig') as f:
                    next(f, None)
                    for line in f:
                        key = line.strip().split(',')[0].strip()
                        if key:
                            existing_targets_artnr.add(key)
        # Generate export, partlist, and tree for all articles in the list
        write_header = not article_list_path.exists() or article_list_path.stat().st_size == 0
        for idx, artnr in enumerate(artnr_list):
            if not check_utf8_file(artikelstamm_path):
                return JSONResponse(status_code=500, content={"status": "error", "message": f"artikelstamm not valid UTF-8: {artikelstamm_path}"})
            # Find the article in artikelstamm
            found = False
            with open(artikelstamm_path, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    if str(row.get('artnr', '')).strip() == str(artnr).strip():
                        artbez1 = row.get('artbez1', '')
                        zeichnr = row.get('zeichnr', '')
                        if artnr in existing_targets_artnr:
                            return JSONResponse(status_code=200, content={"status": "success", "message": f"Article {artnr} already exists in existing articles ({existing_articles_target.upper()}). Nothing appended."})
                        if artnr not in existing_artnr:
                            with open(article_list_path, 'a', encoding='utf-8') as out:
                                if write_header:
                                    out.write('artnr;artbez1;zeichnr\n')
                                    write_header = False
                                out.write(f"{artnr};{artbez1};{zeichnr}\n")
                            existing_artnr.add(artnr)
                        found = True
                        break
            if not found:
                return JSONResponse(status_code=404, content={"status": "error", "message": f"Article {artnr} not found in artikelstamm."})

            # Generate partlist and tree for this article (creation mode, isolated files)
            # Always reset files for the first article, then append for others
            reset_files = (idx == 0)
            process_module_structure(
                str(artnr),
                str(artikelstamm_path),
                str(stuecklistenstamm_path),
                str(article_list_path),
                str(partlist_path),
                mode = "article",
                reset_files=reset_files
            )
        # Return download links (point to creation mode files if needed)
        return {
            "status": "success",
            "message": f"Generated export and partlist for {len(artnr_list)} articles (article mode).",
            "creation_article_list": str(article_list_path.name),
            "creation_partlist": str(partlist_path.name),
            "creation_partlist_tree": str(partlist_tree_path.name)
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})
    
    
@app.post("/reset-article-list")
def reset_article_list():
    """Truncate article-mode list files and write only the headers."""
    base = Path(__file__).parent.parent
    _reset_mode_cache_files(base, "article")
    return {"status": "success", "message": "Article list reset."}

@app.get("/article-list-preview")
def article_list_preview():
    """Return the current contents of article_list_article_mode.csv as JSON."""
    try:
        base = Path(__file__).parent.parent
        article_list_creation_mode_path = _cache_paths(base, "article")["article_list"]
        with open(article_list_creation_mode_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            rows = list(reader)
        return {"status": "success", "rows": rows}
    except Exception as e:
        return {"status": "error", "message": str(e), "rows": []}

@app.post("/remove-article-from-list")
def remove_article_from_list(data: dict = Body(...)):
    """Remove an article by artnr from article_list_article_mode.csv."""
    artnr = str(data.get("artnr", "")).strip()
    if not artnr:
        return {"status": "error", "message": "No artnr provided."}
    base = Path(__file__).parent.parent
    article_list_creation_mode_path = _cache_paths(base, "article")["article_list"]
    try:
        # Read all rows except the one to remove
        with open(article_list_creation_mode_path, 'r', encoding='utf-8-sig') as f:
            rows = list(csv.DictReader(f, delimiter=';'))
        new_rows = [row for row in rows if str(row.get("artnr", "")).strip() != artnr]
        # Write back, preserving header
        with open(article_list_creation_mode_path, 'w', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=["artnr", "artbez1", "zeichnr"], delimiter=';')
            writer.writeheader()
            for row in new_rows:
                writer.writerow(row)
        return {"status": "success", "message": f"Removed article {artnr} from list."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ─── Creation Session State ────────────────────────────────────────────────────


_creation_session: dict | None = None

# --- Hierarchical Structure Endpoints ---

@app.get("/api/creation/structure")
def api_creation_structure():
	"""Return available classes from Struktur.yaml"""
	config_dir = Path(__file__).parent.parent / "config"
	creator = ArticleCreator(config_dir)
	classes = creator.get_classes()
	return {"status": "ok", "classes": classes}


@app.get("/api/creation/groups")
def api_creation_groups(class_name: str = Query(...)):
	"""Return available groups for a given class"""
	config_dir = Path(__file__).parent.parent / "config"
	creator = ArticleCreator(config_dir)
	groups = creator.get_groups_for_class(class_name)
	if not groups:
		return {"status": "error", "message": f"No groups found for class '{class_name}'", "groups": []}
	return {"status": "ok", "groups": groups}


@app.get("/api/creation/types")
def api_creation_types(group_name: str = Query(...)):
	"""Return available types for a given group"""
	config_dir = Path(__file__).parent.parent / "config"
	creator = ArticleCreator(config_dir)
	types = creator.get_types_for_group(group_name)
	if not types:
		return {"status": "error", "message": f"No types found for group '{group_name}'", "types": []}
	# Check if this group has sub-types (e.g., Teileartikel)
	has_subtypes = group_name == "Teileartikel" and len(types) > 0
	return {"status": "ok", "types": types, "has_subtypes": has_subtypes}


@app.get("/api/creation/allowed-children")
def api_creation_allowed_children(parent_type: str = Query(...)):
	"""Return allowed child types for a parent article type"""
	config_dir = Path(__file__).parent.parent / "config"
	creator = ArticleCreator(config_dir)
	allowed = creator.get_allowed_children(parent_type)
	if not allowed:
		return {"status": "info", "message": f"No child types defined for '{parent_type}'", "allowed_types": []}
	return {"status": "ok", "allowed_types": allowed}


@app.post("/api/creation/select-group")
def api_creation_select_group(data: dict = Body(...)):
	"""Select and remember current class and group for article creation"""
	class_name = str(data.get("class", "")).strip()
	group_name = str(data.get("group", "")).strip()

	if not class_name or not group_name:
		return {"status": "error", "message": "class and group parameters required"}

	config_dir = Path(__file__).parent.parent / "config"
	creator = ArticleCreator(config_dir)

	# Validate class and group exist
	classes = creator.get_classes()
	if class_name not in classes:
		return {"status": "error", "message": f"Class '{class_name}' not found"}

	groups = creator.get_groups_for_class(class_name)
	if group_name not in groups:
		return {"status": "error", "message": f"Group '{group_name}' not found in class '{class_name}'"}

	# Get group defaults
	group_defaults = creator.get_group_defaults(group_name)

	return {
		"status": "success",
		"class": class_name,
		"group": group_name,
		"defaults": group_defaults
	}


# ─── Creation Session State ────────────────────────────────────────────────────

# --- Creation Session CSV Utilities ---
def _cs_append_article_list(cache_dir: Path, artnr: str, artbez1: str, zeichnr: str):
    """Append a row to article_list_creation_mode.csv in the cache directory."""
    article_list_path = cache_dir / "article_list_creation_mode.csv"
    write_header = not article_list_path.exists() or article_list_path.stat().st_size == 0
    with open(article_list_path, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        if write_header:
            writer.writerow(["artnr", "artbez1", "zeichnr"])
        writer.writerow([artnr, artbez1, zeichnr])

def _cs_flush_to_csv(session: dict, cache_dir: Path):
    """Write the session state to creation_session.csv and partlist_creation_mode.csv in the cache directory."""
    # Save session to creation_session.csv
    session_path = cache_dir / "creation_session.csv"
    session_path.parent.mkdir(parents=True, exist_ok=True)
    with open(session_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["type","artnr","artbez1","zeichnr","depth","posnr","menge","stulinr","tree_depth","tree_artnr","tree_artbez1","tree_zeichnr"])
        for a in session.get("articles", []):
            writer.writerow([
                a.get("type","article"), a.get("artnr",""), a.get("artbez1",""), a.get("zeichnr",""), a.get("depth",0), "", "", "", "", "", "", ""
            ])
        for p in session.get("partlist_rows", []):
            writer.writerow([
                "", p.get("artnr",""), p.get("artbez1",""), "", "", p.get("posnr",""), p.get("menge",""), p.get("stulinr",""), "", "", "", ""
            ])
        for t in session.get("tree_lines", []):
            writer.writerow([
                "", "", "", "", "", "", "", "", t.get("depth",0), t.get("artnr",""), t.get("artbez1",""), t.get("zeichnr","")
            ])
    # Uniform partlist file for creation mode
    partlist_path = cache_dir / "partlist_creation_mode.csv"
    with open(partlist_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["stulinr","posnr","menge","artnr","artbez1"])
        for p in session.get("partlist_rows", []):
            writer.writerow([
                p.get("stulinr",""), p.get("posnr",""), p.get("menge",""), p.get("artnr",""), p.get("artbez1","")
            ])



# --- Creation Session CSV Storage ---
def _cs_cache_dir() -> Path:
    return Path(__file__).parent.parent / "data" / "processed" / "csv" / "cache"

def _cs_session_csv_path() -> Path:
    return _cs_cache_dir() / "creation_session.csv"

def _cs_new() -> dict:
    return {
        "is_active": False,
        "is_complete": False,
        "current_class": None,
        "current_group": None,
        "stack": [],
        "articles": [],
        "partlist_rows": [],
        "tree_lines": [],
    }

def _cs_save(session: dict) -> None:
    """Save the session to a single CSV file (creation_session.csv) and update partlist_creation_mode.csv for uniformity."""
    path = _cs_session_csv_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    # Write all articles, partlist, and tree info in one CSV
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["type","artnr","artbez1","zeichnr","depth","posnr","menge","stulinr","tree_depth","tree_artnr","tree_artbez1","tree_zeichnr"])
        for a in session.get("articles", []):
            writer.writerow([
                a.get("type","article"), a.get("artnr",""), a.get("artbez1",""), a.get("zeichnr",""), a.get("depth",0), "", "", "", "", "", "", ""
            ])
        for p in session.get("partlist_rows", []):
            writer.writerow([
                "", p.get("artnr",""), p.get("artbez1",""), "", "", p.get("posnr",""), p.get("menge",""), p.get("stulinr",""), "", "", "", ""
            ])
        for t in session.get("tree_lines", []):
            writer.writerow([
                "", "", "", "", "", "", "", "", t.get("depth",0), t.get("artnr",""), t.get("artbez1",""), t.get("zeichnr","")
            ])

    # --- Uniform partlist file for creation mode ---
    partlist_path = _cs_cache_dir() / "partlist_creation_mode.csv"
    with open(partlist_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["stulinr","posnr","menge","artnr","artbez1"])
        for p in session.get("partlist_rows", []):
            writer.writerow([
                p.get("stulinr",""), p.get("posnr",""), p.get("menge",""), p.get("artnr",""), p.get("artbez1","")
            ])

def _cs_load() -> dict:
    """Load the session from the single CSV file (creation_session.csv)."""
    path = _cs_session_csv_path()
    session = _cs_new()
    if not path.exists():
        return session
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            # Article row
            if row["type"]:
                session["articles"].append({
                    "type": row["type"],
                    "artnr": row["artnr"],
                    "artbez1": row["artbez1"],
                    "zeichnr": row["zeichnr"],
                    "depth": int(row["depth"] or 0)
                })
            # Partlist row
            elif row["posnr"]:
                session["partlist_rows"].append({
                    "stulinr": row["stulinr"],
                    "posnr": int(row["posnr"] or 0),
                    "menge": int(row["menge"] or 0),
                    "artnr": row["artnr"],
                    "artbez1": row["artbez1"]
                })
            # Tree line
            elif row["tree_artnr"]:
                session["tree_lines"].append({
                    "depth": int(row["tree_depth"] or 0),
                    "artnr": row["tree_artnr"],
                    "artbez1": row["tree_artbez1"],
                    "zeichnr": row["tree_zeichnr"]
                })
    # Heuristics for is_active/is_complete/stack
    session["is_active"] = not session["is_complete"] and bool(session["stack"] or session["articles"])
    return session

def _cs_tree_text(session: dict) -> str:
    lines = []
    for item in session.get("tree_lines", []):
        depth = item.get("depth", 0)
        artnr = item.get("artnr", "")
        artbez1 = item.get("artbez1", "")
        zeichnr = item.get("zeichnr", "")
        indent = "  " * depth
        zeichnr_part = f" | {zeichnr}" if zeichnr else ""
        lines.append(f"{indent}{artnr} | {artbez1}{zeichnr_part}")
    return "\n".join(lines)


def _cs_do_create_article(creator: "ArticleCreator", type_name: str, input_data: dict, cache_dir: Path, template=None) -> dict:
    """Create + save YAML cache for one article. Returns article field dict."""
    article = creator.create_article(type_name, input_data, template=template)
    artnr = input_data.get("artnr") or input_data.get("modnr") or "article"
    cache_path = cache_dir / f"{type_name}_{artnr}.yaml"
    creator.save_article_cache(article, cache_path)
    return article


@app.get("/api/creation/state")
def api_creation_state():
    """Return the current creation session state."""
    session = _cs_load()
    if not session:
        empty = _cs_new()
        return {**empty, "tree_text": ""}
    return {**session, "tree_text": _cs_tree_text(session)}


@app.post("/api/creation/start")
def api_creation_start(data: dict = Body(...)):
    """Start a new creation session with a root article/module."""
    creator = _load_article_creator()
    type_name = data.get("type")
    class_name = data.get("class")
    group_name = data.get("group")

    if not type_name:
        return {"status": "error", "message": "No article type provided."}
    tpl = _get_effective_template(creator, type_name, group_name)
    if not tpl:
        return {"status": "error", "message": f"Template '{type_name}' not found."}

    input_data = {k: v for k, v in data.items() if k not in ("type", "is_module", "class", "group")}
    cache_dir = _cs_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    try:
        _cs_do_create_article(creator, type_name, input_data, cache_dir, template=tpl)
    except Exception as e:
        return {"status": "error", "message": str(e)}
    artnr = input_data.get("artnr") or input_data.get("modnr") or "root_article"
    artbez1 = input_data.get("artbez1", "")
    zeichnr = input_data.get("zeichnr", "")
    session = _cs_new()
    session["is_active"] = True
    session["current_class"] = class_name
    session["current_group"] = group_name
    session["stack"] = [{"artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr, "pos_counter": 1}]
    session["articles"] = [{"artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr, "type": type_name, "depth": 0}]
    session["tree_lines"] = [{"depth": 0, "artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr}]
    _cs_save(session)
    _cs_append_article_list(cache_dir, artnr, artbez1, zeichnr)
    _cs_flush_to_csv(session, cache_dir)
    return {"status": "success", "artnr": artnr, "state": {**session, "tree_text": _cs_tree_text(session)}}


@app.post("/api/creation/add-child")
def api_creation_add_child(data: dict = Body(...)):
    """Add a child article or module to the current parent."""
    session = _cs_load()
    if not session or not session.get("is_active"):
        return {"status": "error", "message": "No active creation session. Start a root article first."}
    stack = session.get("stack", [])
    if not stack:
        return {"status": "error", "message": "No parent in stack."}
    creator = _load_article_creator()
    type_name = data.get("type")
    group_name = data.get("group")
    is_module = bool(data.get("is_module", False))
    if not type_name:
        return {"status": "error", "message": "No article type provided."}
    tpl = _get_effective_template(creator, type_name, group_name)
    if not tpl:
        return {"status": "error", "message": f"Template '{type_name}' not found."}

    input_data = {k: v for k, v in data.items() if k not in ("type", "is_module", "group")}
    cache_dir = _cs_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    try:
        _cs_do_create_article(creator, type_name, input_data, cache_dir, template=tpl)
    except Exception as e:
        return {"status": "error", "message": str(e)}
    artnr = input_data.get("artnr") or input_data.get("modnr") or "article"
    artbez1 = input_data.get("artbez1", "")
    zeichnr = input_data.get("zeichnr", "")
    parent = stack[-1]
    depth = len(stack)
    # Partlist row
    pos_counter = parent["pos_counter"]
    session["partlist_rows"].append({
        "stulinr": parent["artnr"],
        "posnr": pos_counter,
        "menge": 1,
        "artnr": artnr,
        "artbez1": artbez1,
    })
    session["stack"][-1]["pos_counter"] = pos_counter + 1
    # Register article
    session["articles"].append({"artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr, "type": type_name, "depth": depth})
    session["tree_lines"].append({"depth": depth, "artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr})
    # If it's a module, push it onto the stack so subsequent children attach to it
    if is_module:
        session["stack"].append({"artnr": artnr, "artbez1": artbez1, "zeichnr": zeichnr, "pos_counter": 1})
    _cs_save(session)
    _cs_append_article_list(cache_dir, artnr, artbez1, zeichnr)
    _cs_flush_to_csv(session, cache_dir)
    return {"status": "success", "artnr": artnr, "is_module": is_module, "state": {**session, "tree_text": _cs_tree_text(session)}}


@app.post("/api/creation/finish-module")
def api_creation_finish_module():
    """Finish the current module and go up one level. If at root, mark session complete."""
    session = _cs_load()
    if not session or not session.get("is_active"):
        return {"status": "error", "message": "No active creation session."}
    stack = session.get("stack", [])
    if len(stack) <= 1:
        session["is_active"] = False
        session["is_complete"] = True
        _cs_save(session)
        _cs_flush_to_csv(session, _cs_cache_dir())
        return {"status": "complete", "message": "Root module finished. Session complete.", "state": {**session, "tree_text": _cs_tree_text(session)}}
    session["stack"].pop()
    _cs_save(session)
    return {"status": "success", "state": {**session, "tree_text": _cs_tree_text(session)}}


@app.post("/api/creation/reset")
def api_creation_reset():
    """Reset the creation session and clear related files."""
    global _creation_session
    _creation_session = None
    cache_dir = _cs_cache_dir()
    for fname in [
        "creation_session.json",
        "article_list_creation_mode.csv",
        "partlist_creation_mode.csv",
        "partlist_creation_mode_tree.txt",
    ]:
        p = cache_dir / fname
        if p.exists():
            p.unlink()
    return {"status": "success", "message": "Creation session reset."}


# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/article-types")
def api_article_types():
    """
    Returns a list of available article types/templates from article_templates.yaml
    """
    config_dir = Path(__file__).parent.parent / "config"
    creator = ArticleCreator(config_dir)
    types = creator.list_templates()
    return {"types": types}

@app.get("/api/article-fields")
def api_article_fields(type: str, group: str | None = None):
    """
    Returns the fields for a given article type/template.
    """
    creator = _load_article_creator()
    tpl = _get_effective_template(creator, type, group)
    if not tpl:
        return {"status": "error", "fields": []}
    # Return all fields with their properties
    fields = [f.as_dict() for f in tpl.fields]
    return {"status": "ok", "fields": fields}

@app.post("/api/generate-root-article")
def api_generate_root_article(data: dict = Body(...)):
    """
    Creates and saves a root article of the selected type with the provided data.
    """
    creator = _load_article_creator()
    type_name = data.get("type")
    group_name = data.get("group")
    if not type_name:
        return {"status": "error", "message": "No article type provided."}
    tpl = _get_effective_template(creator, type_name, group_name)
    if not tpl:
        return {"status": "error", "message": f"Template '{type_name}' not found."}
    # Remove 'type' from data
    input_data = {k: v for k, v in data.items() if k not in {"type", "group"}}
    try:
        article = creator.create_article(type_name, input_data, template=tpl)
        # Save to cache (YAML for now, could be CSV)
        artnr = input_data.get("artnr") or input_data.get("modnr") or "root_article"
        cache_path = cache_dir_path / f"{type_name}_{artnr}.yaml"
        creator.save_article_cache(article, cache_path)

        # --- Also append to partlist_creation_mode.csv and partlist_creation_mode_tree.txt ---
        partlist_path = get_path("partlist", "creation_mode")
        tree_path = get_path("partlisttree", "creation_mode")

        # Compose values for partlist and tree
        artbez1 = input_data.get("artbez1", "")
        zeichnr = input_data.get("zeichnr", "")
        # For partlist: stulinr;posnr;menge;artnr;artbez1
        partlist_row = {
            "stulinr": artnr,
            "posnr": "1",
            "menge": "1",
            "artnr": artnr,
            "artbez1": artbez1
        }
        # Write header if file does not exist
        write_header = not partlist_path.exists() or partlist_path.stat().st_size == 0
        with open(partlist_path, "a", encoding="utf-8") as f:
            if write_header:
                f.write("stulinr;posnr;menge;artnr;artbez1\n")
            f.write(f"{partlist_row['stulinr']};{partlist_row['posnr']};{partlist_row['menge']};{partlist_row['artnr']};{partlist_row['artbez1']}\n")

        # For tree: artnr, artbez1, zeichnr
        with open(tree_path, "a", encoding="utf-8") as f:
            f.write(f"{artnr}, {artbez1}, {zeichnr}\n")

        return {"status": "success", "message": f"{type_name} created and saved as {cache_path.name}. Also added to partlist and tree.", "article": article}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/check-article-exists")
def check_article_exists(artnr: str, mode: str = Query("article")):
    """
    Check if an article exists in Majesty data (artikelstamm CSV) or in creation cache if mode=creation.
    """
    base = Path(__file__).parent.parent
    if mode == "creation":
        # Check active session articles first (in-memory)
        active_session = _cs_load()
        if active_session:
            for art in active_session.get("articles", []):
                if str(art.get("artnr", "")).strip() == str(artnr).strip():
                    return {"status": "found_in_creation", "artnr": artnr, "data": art}
        # Fallback: check creation cache CSV
        article_list_creation_mode_path = base / "data" / "processed" / "csv" / "cache" / "article_list_creation_mode.csv"
        if article_list_creation_mode_path.exists():
            with open(article_list_creation_mode_path, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    if str(row.get("artnr", "")).strip() == str(artnr).strip():
                        return {"status": "found_in_creation", "artnr": artnr, "data": row}
    # Always check Majesty data as well
    if not artikelstamm_path.exists():
        return {"status": "error", "message": "Majesty artikelstamm not found."}
    with open(artikelstamm_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if str(row.get("artnr", "")).strip() == str(artnr).strip():
                return {"status": "found", "artnr": artnr, "data": row}
    return {"status": "not_found", "artnr": artnr}

# Live field-by-field YAML cache endpoint
@app.post("/api/cache-article-field")
def cache_article_field(data: dict = Body(...)):
    """
    Cache a single field value for an article/module to a YAML file.
    Expects: {"cache_id": str, "field": str, "value": any}
    """
    cache_id = str(data.get("cache_id") or "").strip()
    field = str(data.get("field") or "").strip()
    value = data.get("value")
    if not cache_id or not field:
        return {"status": "error", "message": "cache_id and field required"}
    base = Path(__file__).parent.parent
    cache_dir = base / "data" / "processed" / "csv" / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{cache_id}.yaml"
    # Load existing cache
    cache_data = {}
    if cache_file.exists():
        with open(cache_file, "r", encoding="utf-8") as f:
            try:
                cache_data = yaml.safe_load(f) or {}
            except Exception:
                cache_data = {}
    cache_data[field] = value
    with open(cache_file, "w", encoding="utf-8") as f:
        yaml.safe_dump(cache_data, f, allow_unicode=True)
    return {"status": "ok", "cache_id": cache_id, "field": field, "value": value}


@app.get("/api/validate-artikelbezeichnungen")
def api_validate_artikelbezeichnungen():
    base = Path(__file__).parent.parent
    try:
        return load_article_validation_queue(base)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})


@app.post("/api/validate-artikelbezeichnungen/save")
def api_validate_artikelbezeichnungen_save(data: dict = Body(...)):
    base = Path(__file__).parent.parent
    artnr = str(data.get("artnr") or "").strip()
    updates = data.get("updates") or {}
    try:
        return save_article_validation_item(base, artnr, updates)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})


@app.post("/api/validate-artikelbezeichnungen/overwrite")
def api_validate_artikelbezeichnungen_overwrite():
    base = Path(__file__).parent.parent
    try:
        return overwrite_bezeichnungen_from_config(base)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})


@app.get('/api/validate/groups')
def api_validate_groups():
    try:
        base = Path(__file__).parent.parent
        return load_article_group_validation_queue(base)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})


@app.post('/api/validate/groups/save')
def api_validate_groups_save(data: dict = Body(...)):
    artnr = str((data or {}).get('artnr', '')).strip()
    if not artnr:
        return {"status": "error", "message": "artnr required"}
    updates = {
        'hauptgruppe': (data or {}).get('hauptgruppe', ''),
        'untergruppe': (data or {}).get('untergruppe', ''),
        'spezifikation': (data or {}).get('spezifikation', ''),
        'bezeichnungselemente': (data or {}).get('bezeichnungselemente', []),
    }
    try:
        base = Path(__file__).parent.parent
        return save_article_group_item(base, artnr, updates)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})


@app.post('/api/validate/groups/add')
def api_validate_groups_add(data: dict = Body(...)):
    """Add a new hauptgruppe / untergruppe / spezifikation into config/groups.yaml."""
    try:
        base = Path(__file__).parent.parent
        hg = str((data or {}).get('hauptgruppe', '') or '').strip()
        ug = str((data or {}).get('untergruppe', '') or '').strip()
        spec = str((data or {}).get('spezifikation', '') or '').strip()
        if not hg:
            return {"status": "error", "message": "hauptgruppe is required"}
        return add_group_entry(base, hg, ug, spec)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})

@app.post('/api/validate/groups/resolve-bezeichnungselemente')
def api_validate_groups_resolve_bezeichnungselemente(data: dict = Body(...)):
    """Resolve Bezeichnungselemente for an article given a custom group path."""
    try:
        from etl.transform import getBezeichnungselemente
        artnr = str((data or {}).get('artnr', '')).strip()
        hg = str((data or {}).get('hauptgruppe', '')).strip()
        ug = str((data or {}).get('untergruppe', '')).strip()
        spec = str((data or {}).get('spezifikation', '')).strip()
        
        if not artnr:
            return {"status": "error", "message": "artnr is required"}
        
        elements = getBezeichnungselemente(
            artnr,
            group_path={
                'hauptgruppe': hg,
                'untergruppe': ug,
                'spezifikation': spec,
            },
        ) or []
        return {"status": "ok", "bezeichnungselemente": elements}
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(exc)})

@app.get("/api/search-lieferant")
def search_lieferant(q: str = Query("", min_length=1)):
    """
    Search lieferanten_mapping.csv for supplier by name or ifas_nummer. Returns list of dicts with ifas_nummer and name.
    """
    base = Path(__file__).parent.parent
    mapping_path = base / "config" / "lieferanten_mapping.csv"
    results = []
    if not mapping_path.exists():
        return JSONResponse(status_code=404, content={"error": "lieferanten_mapping.csv not found"})
    try:
        with open(mapping_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=';' if ';' in f.read(1024) else ',')
            f.seek(0)
            for row in reader:
                ifas_nummer = str(row.get("ifas_nummer", "")).strip()
                name = str(row.get("name", "")).strip()
                # Search in ifas_nummer or name (case-insensitive)
                if q.lower() in ifas_nummer.lower() or q.lower() in name.lower():
                    results.append({"ifas_nummer": ifas_nummer, "name": name})
                    if len(results) >= 20:
                        break
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
    return results

@app.get("/download-partlist")
def download_partlist(artnr: str = Query(..., min_length=1), mode: str = Query("creation")):
    """Download the partlist CSV for the current session (creation mode)."""
    base = Path(__file__).parent.parent
    partlist_path = _cache_paths(base, mode)["partlist"]
    if not partlist_path.exists():
        return JSONResponse(status_code=404, content={"error": f"partlist.csv not found for {artnr} (mode={mode})"})
    try:
        return FileResponse(str(partlist_path), filename=f"partlist_{artnr}.csv", media_type="text/csv")
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# --- Download Group Excel for Creation Session ---
@app.get("/download-group-excel")
def download_group_excel(artnr: str = Query(..., min_length=1), mode: str = Query("creation")):
    """Download a group Excel (XLSX) for the current creation session."""
    base = Path(__file__).parent.parent
    partlist_csv = _cache_paths(base, mode)["partlist"]
    template_xlsx = base / "data" / "raw" / "templates" / "VorlageStücklisteV1.xlsx"
    output_xlsx = base / "data" / "processed" / "csv" / "cache" / f"group_export_{artnr}.xlsx"
    if not partlist_csv.exists() or not template_xlsx.exists():
        return JSONResponse(status_code=404, content={"error": "partlist.csv or template not found."})
    try:
        create_partlist_excel_from_template(partlist_csv, template_xlsx, output_xlsx)
        return FileResponse(str(output_xlsx), filename=f"group_export_{artnr}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/add-stulinr-from-stuecklisten")
def add_stulinr_from_stuecklisten(partlist_env: str = Body("PROD")):
    """
    Add all stulinr from the sheet stuecklisten to the existing partlists cache (test/prod).
    """
    stuecklisten_path = get_path("stuecklistenstamm")
    stulinr_set = set()
    with open(stuecklisten_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            stulinr = row.get('stulinr', '').strip()
            if stulinr:
                stulinr_set.add(stulinr)

    # Write to processed cache existing partlists
    cache_dir = BASE_DIR / "data" / "processed" / "cache" / "existing"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"partlists_{partlist_env}.csv"
    # Append new stulinr, avoid duplicates
    existing = set()
    if cache_file.exists():
        with open(cache_file, encoding="utf-8") as f:
            for line in f:
                val = line.strip().split(';')[0]
                if val:
                    existing.add(val)
    with open(cache_file, 'a', encoding='utf-8') as f:
        for stulinr in stulinr_set:
            if stulinr not in existing:
                f.write(f"{stulinr};\n")
    return {"status": "ok", "added": list(stulinr_set - existing)}
@app.post("/upload-existing-partlist")
async def upload_existing_partlist(existing_partlist_file: UploadFile = File(...), partlist_env: str = Form(...)):
    """
    Upload an existing partlist CSV (test/prod), extract all stulinr, and add to processed cache existing partlists.
    """
    # Save uploaded file to temp
    temp_path = BASE_DIR / "data" / "uploaded_files" / f"uploaded_existing_partlist_{partlist_env}.csv"
    with open(temp_path, "wb") as buffer:
        buffer.write(await existing_partlist_file.read())

    # Extract stulinr from uploaded CSV
    stulinr_set = set()
    with open(temp_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            stulinr = row.get('stulinr', '').strip()
            if stulinr:
                stulinr_set.add(stulinr)

    # Write to processed cache existing partlists
    cache_dir = BASE_DIR / "data" / "processed" / "cache" / "existing"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"partlists_{partlist_env}.csv"
    # Append new stulinr, avoid duplicates
    existing = set()
    if cache_file.exists():
        with open(cache_file, encoding="utf-8") as f:
            for line in f:
                val = line.strip().split(';')[0]
                if val:
                    existing.add(val)
    with open(cache_file, 'a', encoding='utf-8') as f:
        for stulinr in stulinr_set:
            if stulinr not in existing:
                f.write(f"{stulinr};\n")
    return {"status": "ok", "added": list(stulinr_set - existing)}
# --- Document XLSX Export Endpoint ---
from etl.load import export_docs_xlsx

@app.get("/download-docs-xlsx")
def download_docs_xlsx(mode: str = Query("module")):
    """
    Download an XLSX file with doc file info for all articles with zeichnungsnummer in the current article list.
    """
    article_list_path = _cache_paths(BASE_DIR, mode)["article_list"]
    if not article_list_path.exists():
        return JSONResponse(status_code=404, content={"error": "article_list.csv not found. Generate module data first."})
    articles = _read_article_list_rows(article_list_path)
    if not articles:
        return JSONResponse(status_code=404, content={"error": f"No articles found in {article_list_path.name}. Generate {mode} data first."})

    docs_cache_path = sheets_output_dir / "DMSDocuments.csv"
    docs_cache_result = build_docs_download_cache_csv(
        article_list_path=article_list_path,
        output_csv_path=docs_cache_path,
        articles=articles,
    )
    if docs_cache_result.get("rows_written", 0) == 0:
        return JSONResponse(status_code=404, content={
            "error": f"No DMS document rows could be generated for mode={mode}.",
            "docs_cache_path": str(docs_cache_path),
        })

    xlsx_path = export_docs_xlsx(articles)
    return FileResponse(
        str(xlsx_path),
        filename="docs_export.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/api/generate-article-number")
def api_generate_article_number(data: dict = Body(...)):
    """
    Generate the next available article number in XX.YY.ZZZZ format.

    Input: {"prefix": ""} (optional prefix)
    - Empty: lowest unused number (00.00.0000+)
    - "10": lowest starting with 10 (10.00.0000+)
    - "10.20": lowest starting with 10.20 (10.20.0000+)
    - "10.20.5678": that number if available, else next

    Returns: {"status": "ok", "number": "XX.YY.ZZZZ"} or {"status": "error", "message": "..."}
    """
    import logging
    try:
        from etl.Nummer_vergeben import generate_article_number

        config_dir = BASE_DIR
        prefix = data.get("prefix", "").strip()

        logging.info(f"[API] Generating article number with prefix='{prefix}'")
        number = generate_article_number(config_dir, prefix)
        logging.info(f"[API] Successfully generated: {number}")
        return {"status": "ok", "number": number}
    except ValueError as e:
        logging.warning(f"[API] ValueError during generation: {e}")
        return {"status": "error", "message": str(e)}
    except Exception as e:
        logging.error(f"[API] Unexpected error during generation: {type(e).__name__}: {e}")
        import traceback
        logging.error(f"[API] Traceback: {traceback.format_exc()}")
        return {"status": "error", "message": f"Error: {str(e)}"}


@app.get("/api/generated-numbers")
def api_get_generated_numbers():
    """
    Get list of all previously generated article numbers.

    Returns: {"status": "ok", "numbers": [...], "count": N}
    """
    try:
        from etl.Nummer_vergeben import _load_generated_numbers

        config_dir = BASE_DIR
        generated = _load_generated_numbers(config_dir)

        return {
            "status": "ok",
            "numbers": sorted(list(generated)),
            "count": len(generated)
        }
    except Exception as e:
        logging.error(f"Error loading generated numbers: {e}")
        return {"status": "error", "message": str(e)}

    